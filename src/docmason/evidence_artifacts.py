"""Phase 3 spreadsheet and multimodal artifact compilers."""

from __future__ import annotations

import os
import re
from collections import Counter, defaultdict
from importlib import import_module
from pathlib import Path
from statistics import mean, median
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries  # type: ignore[import-untyped]

from .artifacts import (
    artifact_graph_promoted,
    artifact_locator_aliases,
    artifact_title_from_text,
    deduplicate_strings,
    normalize_bbox,
    stable_artifact_id,
)
from .project import read_json, write_json

TIME_AXIS_TOKENS = ("date", "day", "week", "month", "quarter", "year", "timeline", "period")
METRIC_TOKENS = ("kpi", "metric", "value", "score", "cost", "budget", "count", "rate", "risk")
DIMENSION_TOKENS = (
    "segment",
    "region",
    "category",
    "channel",
    "owner",
    "team",
    "source",
    "country",
)
ROLE_HINT_PATTERNS = {
    "architecture-like": ("architecture", "platform", "service", "system", "integration"),
    "flow-like": ("flow", "workflow", "process", "step", "stage", "journey"),
    "ui-like": ("screen", "ui", "interface", "dashboard", "login", "button"),
    "comparison-like": ("compare", "comparison", "versus", "vs", "before", "after"),
    "roadmap-like": ("roadmap", "timeline", "milestone", "phase", "quarter"),
    "kpi-like": ("kpi", "metric", "scorecard", "performance", "variance"),
    "overview-like": ("overview", "summary", "at a glance"),
    "appendix-like": ("appendix", "backup"),
}
MONTH_TOKENS = (
    "jan",
    "feb",
    "mar",
    "apr",
    "may",
    "jun",
    "jul",
    "aug",
    "sep",
    "sept",
    "oct",
    "nov",
    "dec",
    "january",
    "february",
    "march",
    "april",
    "june",
    "july",
    "august",
    "september",
    "october",
    "november",
    "december",
)
CHART_TOKENS = (
    "chart",
    "graph",
    "trend",
    "forecast",
    "actual",
    "budget",
    "target",
    "variance",
    "revenue",
    "sales",
    "growth",
    "performance",
    "q1",
    "q2",
    "q3",
    "q4",
)
CAPTION_PREFIX_PATTERN = re.compile(
    r"^(figure|fig\.|table|chart|diagram|exhibit)\s+[A-Za-z0-9.-]+\s*[:.-]?\s+",
    re.IGNORECASE,
)
HEADING_NUMBER_PATTERN = re.compile(r"^(?:\d+(?:\.\d+){0,5}|[A-Z])(?:[.)-])?\s+\S")
ORDERED_STEP_PATTERN = re.compile(r"^\s*(\d+)(?:[.)-])\s+")
BULLET_STEP_PATTERN = re.compile(r"^\s*[-*•]\s+")


def _utc_now() -> str:
    from datetime import UTC, datetime

    return datetime.now(tz=UTC).isoformat().replace("+00:00", "Z")


def _sanitize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = text.encode("utf-8", errors="replace").decode("utf-8")
    return "\n".join(line.rstrip() for line in text.splitlines()).strip()


def _compact_text(value: Any, *, limit: int = 180) -> str:
    compact = re.sub(r"\s+", " ", _sanitize_text(value)).strip()
    if len(compact) <= limit:
        return compact
    truncated = compact[:limit].rsplit(" ", 1)[0].strip()
    return (truncated or compact[:limit]).rstrip(" ,;:.") + "..."


def _slugify_text(value: Any, *, limit: int = 48) -> str:
    compact = re.sub(r"[^0-9A-Za-z]+", "-", _sanitize_text(value).lower()).strip("-")
    return compact[:limit] or "untitled"


def _safe_float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _bbox_from_grid(
    *,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
) -> dict[str, float]:
    return {
        "x0": float(min_col - 1),
        "y0": float(min_row - 1),
        "x1": float(max_col),
        "y1": float(max_row),
    }


def _empty_artifact_index(source_id: str) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "artifact_type": "artifact-index",
        "generated_at": _utc_now(),
        "source_id": source_id,
        "artifacts": [],
    }


def write_empty_artifact_index(source_dir: Path, *, source_id: str) -> dict[str, Any]:
    payload = _empty_artifact_index(source_id)
    write_json(source_dir / "artifact_index.json", payload)
    return payload


def _write_artifact_index(
    source_dir: Path,
    *,
    source_id: str,
    artifacts: list[dict[str, Any]],
) -> dict[str, Any]:
    payload = _empty_artifact_index(source_id)
    payload["artifacts"] = artifacts
    write_json(source_dir / "artifact_index.json", payload)
    return payload


def _role_hints_from_text(texts: list[str], *, artifact_types: list[str]) -> list[str]:
    haystack = " ".join(value.lower() for value in texts if value)
    hints = [
        hint
        for hint, tokens in ROLE_HINT_PATTERNS.items()
        if any(token in haystack for token in tokens)
    ]
    if "chart" in artifact_types and "kpi-like" not in hints:
        hints.append("kpi-like")
    if "picture" in artifact_types and any(
        token in haystack for token in ROLE_HINT_PATTERNS["ui-like"]
    ):
        hints.append("ui-like")
    return deduplicate_strings(hints)


def _open_pymupdf() -> Any:
    try:
        pymupdf = import_module("pymupdf")
    except ImportError:  # pragma: no cover - compatibility import
        pymupdf = import_module("fitz")
    _configure_pymupdf_console_output(pymupdf)
    return pymupdf


def _configure_pymupdf_console_output(pymupdf: Any) -> None:
    """Keep PyMuPDF warnings inside DocMason instead of polluting CLI JSON output."""
    try:
        pymupdf.no_recommend_layout()
    except Exception:
        pass
    try:
        pymupdf.TOOLS.mupdf_display_errors(False)
    except Exception:
        pass
    try:
        pymupdf.TOOLS.mupdf_display_warnings(False)
    except Exception:
        pass
    try:
        pymupdf.set_messages(path=os.devnull)
    except Exception:
        pass


def _consume_pymupdf_warnings(pymupdf: Any) -> list[str]:
    """Return deduplicated MuPDF warnings recorded during the current operation."""
    try:
        raw = str(pymupdf.TOOLS.mupdf_warnings() or "")
    except Exception:
        raw = ""
    try:
        pymupdf.TOOLS.reset_mupdf_warnings()
    except Exception:
        pass
    return deduplicate_strings(
        [line.strip() for line in raw.splitlines() if isinstance(line, str) and line.strip()]
    )


def _bbox_from_pdf_rect(value: Any) -> dict[str, float] | None:
    if isinstance(value, dict):
        try:
            return {
                "x0": float(value["x0"]),
                "y0": float(value["y0"]),
                "x1": float(value["x1"]),
                "y1": float(value["y1"]),
            }
        except (KeyError, TypeError, ValueError):
            return None
    if isinstance(value, (list, tuple)) and len(value) == 4:
        try:
            return {
                "x0": float(value[0]),
                "y0": float(value[1]),
                "x1": float(value[2]),
                "y1": float(value[3]),
            }
        except (TypeError, ValueError):
            return None
    for attribute in ("x0", "y0", "x1", "y1"):
        if not hasattr(value, attribute):
            return None
    try:
        return {
            "x0": float(value.x0),
            "y0": float(value.y0),
            "x1": float(value.x1),
            "y1": float(value.y1),
        }
    except (TypeError, ValueError):
        return None


def _bbox_area(bbox: dict[str, float] | None) -> float:
    if not isinstance(bbox, dict):
        return 0.0
    return max(float(bbox["x1"]) - float(bbox["x0"]), 0.0) * max(
        float(bbox["y1"]) - float(bbox["y0"]),
        0.0,
    )


def _expand_bbox(bbox: dict[str, float], *, padding: float) -> dict[str, float]:
    return {
        "x0": float(bbox["x0"]) - padding,
        "y0": float(bbox["y0"]) - padding,
        "x1": float(bbox["x1"]) + padding,
        "y1": float(bbox["y1"]) + padding,
    }


def _bboxes_overlap(
    left: dict[str, float],
    right: dict[str, float],
    *,
    padding: float = 0.0,
) -> bool:
    expanded_left = _expand_bbox(left, padding=padding)
    expanded_right = _expand_bbox(right, padding=padding)
    return not (
        expanded_left["x1"] < expanded_right["x0"]
        or expanded_right["x1"] < expanded_left["x0"]
        or expanded_left["y1"] < expanded_right["y0"]
        or expanded_right["y1"] < expanded_left["y0"]
    )


def _union_bbox(boxes: list[dict[str, float]]) -> dict[str, float] | None:
    if not boxes:
        return None
    return {
        "x0": min(float(box["x0"]) for box in boxes),
        "y0": min(float(box["y0"]) for box in boxes),
        "x1": max(float(box["x1"]) for box in boxes),
        "y1": max(float(box["y1"]) for box in boxes),
    }


def _total_bbox_area(boxes: list[dict[str, float]]) -> float:
    return sum(_bbox_area(box) for box in boxes if isinstance(box, dict))


def _pdf_text_layer_quality(
    *,
    extracted_text: str,
    word_records: list[dict[str, Any]],
    picture_boxes: list[dict[str, float]],
    graphic_boxes: list[dict[str, float]],
    page_area: float,
) -> tuple[str, list[str]]:
    extracted_length = len(_sanitize_text(extracted_text))
    word_count = len(word_records)
    word_text_length = len(
        " ".join(
            str(item.get("text") or "") for item in word_records if isinstance(item, dict)
        ).strip()
    )
    image_area_ratio = min(_total_bbox_area(picture_boxes) / max(page_area, 1.0), 1.0)
    graphic_area_ratio = min(_total_bbox_area(graphic_boxes) / max(page_area, 1.0), 1.0)
    gap_hints: list[str] = []
    if word_count >= 25 or extracted_length >= 160:
        quality = "rich"
    elif word_count >= 5 or extracted_length >= 24 or word_text_length >= 24:
        quality = "weak"
    else:
        quality = "none"
    if image_area_ratio >= 0.82 and word_count <= 2 and extracted_length <= 24:
        gap_hints.append("image-only-page")
    elif image_area_ratio >= 0.55 and quality != "rich":
        gap_hints.append("scanned-page-like")
    if quality == "weak" and max(image_area_ratio, graphic_area_ratio) >= 0.25:
        gap_hints.append("weak-text-layer")
    if max(extracted_length, word_text_length) >= 80 and min(
        extracted_length, word_text_length
    ) <= 0.35 * max(extracted_length, word_text_length):
        gap_hints.append("text-layer-mismatch")
    if max(image_area_ratio, graphic_area_ratio) >= 0.35:
        gap_hints.append("visual-heavy-page")
    return quality, deduplicate_strings(gap_hints)


def _pdf_artifact_semantic_gap_hints(
    *,
    artifact_type: str,
    linked_text: str,
    caption_text: str,
    page_gap_hints: list[str],
) -> list[str]:
    gap_hints: list[str] = []
    if artifact_type == "page-image":
        gap_hints.extend(page_gap_hints)
    if artifact_type in {"chart", "picture", "major-region"} and not (linked_text or caption_text):
        gap_hints.extend(
            hint
            for hint in page_gap_hints
            if hint
            in {"image-only-page", "scanned-page-like", "weak-text-layer", "text-layer-mismatch"}
        )
    if artifact_type == "table" and not linked_text:
        gap_hints.extend(
            hint for hint in page_gap_hints if hint in {"weak-text-layer", "text-layer-mismatch"}
        )
    return deduplicate_strings(gap_hints)


def _pdf_word_records(page: Any) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for item in page.get_text("words"):
        if not isinstance(item, (list, tuple)) or len(item) < 5:
            continue
        text = _sanitize_text(item[4])
        if not text:
            continue
        bbox = _bbox_from_pdf_rect(item[:4])
        if bbox is None:
            continue
        records.append({"bbox": bbox, "text": text})
    return records


def _pdf_word_lines(words: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    grouped: list[list[dict[str, Any]]] = []
    for record in sorted(words, key=lambda item: (item["bbox"]["y0"], item["bbox"]["x0"])):
        if not grouped:
            grouped.append([record])
            continue
        previous_line = grouped[-1]
        previous_center = mean(
            [
                (float(item["bbox"]["y0"]) + float(item["bbox"]["y1"])) / 2.0
                for item in previous_line
            ]
        )
        current_center = (float(record["bbox"]["y0"]) + float(record["bbox"]["y1"])) / 2.0
        if abs(current_center - previous_center) <= 4.5:
            previous_line.append(record)
        else:
            grouped.append([record])
    return [sorted(line, key=lambda item: float(item["bbox"]["x0"])) for line in grouped if line]


def _cells_from_pdf_word_line(
    words: list[dict[str, Any]], *, gap_threshold: float = 32.0
) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    current_words: list[dict[str, Any]] = []
    for word in words:
        if not current_words:
            current_words.append(word)
            continue
        gap = float(word["bbox"]["x0"]) - float(current_words[-1]["bbox"]["x1"])
        if gap <= gap_threshold:
            current_words.append(word)
            continue
        cells.append(
            {
                "text": " ".join(str(item["text"]) for item in current_words).strip(),
                "bbox": _union_bbox([item["bbox"] for item in current_words]),
            }
        )
        current_words = [word]
    if current_words:
        cells.append(
            {
                "text": " ".join(str(item["text"]) for item in current_words).strip(),
                "bbox": _union_bbox([item["bbox"] for item in current_words]),
            }
        )
    return [
        cell
        for cell in cells
        if isinstance(cell.get("bbox"), dict) and isinstance(cell.get("text"), str) and cell["text"]
    ]


def _manual_pdf_table_candidates(
    word_lines: list[list[dict[str, Any]]],
    *,
    excluded_bboxes: list[dict[str, float]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    current_rows: list[list[dict[str, Any]]] = []
    for line in word_lines:
        cells = _cells_from_pdf_word_line(line)
        if len(cells) < 2:
            if current_rows:
                candidates.extend(
                    _table_candidates_from_rows(current_rows, excluded_bboxes=excluded_bboxes)
                )
                current_rows = []
            continue
        if current_rows:
            previous_y = float(current_rows[-1][0]["bbox"]["y0"])
            current_y = float(cells[0]["bbox"]["y0"])
            if abs(current_y - previous_y) > 28.0:
                candidates.extend(
                    _table_candidates_from_rows(current_rows, excluded_bboxes=excluded_bboxes)
                )
                current_rows = []
        current_rows.append(cells)
    if current_rows:
        candidates.extend(
            _table_candidates_from_rows(current_rows, excluded_bboxes=excluded_bboxes)
        )
    return candidates


def _table_candidates_from_rows(
    rows: list[list[dict[str, Any]]],
    *,
    excluded_bboxes: list[dict[str, float]],
) -> list[dict[str, Any]]:
    if len(rows) < 3:
        return []
    column_count = Counter(len(row) for row in rows).most_common(1)[0][0]
    if column_count < 2:
        return []
    aligned_rows = [row for row in rows if len(row) >= column_count]
    if len(aligned_rows) < 3:
        return []
    start_positions: list[list[float]] = [[] for _ in range(column_count)]
    for row in aligned_rows:
        for index, cell in enumerate(row[:column_count]):
            start_positions[index].append(float(cell["bbox"]["x0"]))
    if any((max(values) - min(values)) > 38.0 for values in start_positions if values):
        return []
    bbox = _union_bbox(
        [
            cell["bbox"]
            for row in aligned_rows
            for cell in row[:column_count]
            if isinstance(cell.get("bbox"), dict)
        ]
    )
    if bbox is None:
        return []
    if any(_bboxes_overlap(bbox, excluded_bbox, padding=6.0) for excluded_bbox in excluded_bboxes):
        return []
    preview = [[str(cell["text"]) for cell in row[:column_count]] for row in aligned_rows[:8]]
    return [{"bbox": bbox, "rows": preview}]


def _pdf_text_lines(page_dict: dict[str, Any]) -> list[dict[str, Any]]:
    lines: list[dict[str, Any]] = []
    for block in page_dict.get("blocks", []):
        if not isinstance(block, dict) or int(block.get("type", -1)) != 0:
            continue
        for line in block.get("lines", []):
            if not isinstance(line, dict):
                continue
            spans = [span for span in line.get("spans", []) if isinstance(span, dict)]
            text_parts = [_sanitize_text(span.get("text")) for span in spans]
            text = " ".join(part for part in text_parts if part).strip()
            if not text:
                continue
            sizes: list[float] = []
            flags: list[int] = []
            for span in spans:
                size_value = span.get("size")
                if isinstance(size_value, (int, float)):
                    sizes.append(float(size_value))
                flag_value = span.get("flags")
                if isinstance(flag_value, int):
                    flags.append(flag_value)
            bbox = _bbox_from_pdf_rect(line.get("bbox"))
            if bbox is None:
                continue
            lines.append(
                {
                    "text": text,
                    "bbox": bbox,
                    "font_size": max(sizes) if sizes else 0.0,
                    "font_names": [
                        str(span.get("font"))
                        for span in spans
                        if isinstance(span.get("font"), str) and span.get("font")
                    ],
                    "flags": flags,
                }
            )
    return lines


def _heading_level_from_text(text: str, *, font_size: float, body_size: float) -> int:
    stripped = text.strip()
    numbered = re.match(r"^(\d+(?:\.\d+)*)", stripped)
    if numbered:
        return min(numbered.group(1).count(".") + 1, 6)
    if font_size >= (body_size * 1.75):
        return 1
    if font_size >= (body_size * 1.45):
        return 2
    return 3


def _pdf_heading_candidates(
    page_lines: list[dict[str, Any]],
    *,
    page_height: float,
) -> tuple[list[dict[str, Any]], float]:
    font_sizes = [
        float(line["font_size"])
        for line in page_lines
        if isinstance(line.get("font_size"), (int, float)) and float(line["font_size"]) > 0
    ]
    body_size = median(font_sizes) if font_sizes else 11.0
    candidates: list[dict[str, Any]] = []
    for line in page_lines:
        text = str(line.get("text") or "").strip()
        if not text:
            continue
        word_count = len(text.split())
        if word_count > 14:
            continue
        bbox = line.get("bbox")
        if not isinstance(bbox, dict):
            continue
        font_size = float(line.get("font_size") or body_size)
        score = 0
        if font_size >= (body_size * 1.2):
            score += 2
        if float(bbox["y0"]) <= (page_height * 0.24):
            score += 1
        if HEADING_NUMBER_PATTERN.match(text):
            score += 1
        if text.isupper() and word_count <= 8:
            score += 1
        if text.endswith("."):
            score -= 1
        if score < 2:
            continue
        confidence = "high" if score >= 4 else "medium" if score >= 3 else "low"
        candidates.append(
            {
                "text": text,
                "level": _heading_level_from_text(text, font_size=font_size, body_size=body_size),
                "bbox": bbox,
                "font_size": round(font_size, 2),
                "confidence": confidence,
            }
        )
    return candidates, body_size


def _outline_nodes_from_pdf(document: Any) -> list[dict[str, Any]]:
    try:
        raw_toc = document.get_toc(simple=False)
    except TypeError:  # pragma: no cover - compatibility with older PyMuPDF
        raw_toc = document.get_toc()
    except Exception:  # pragma: no cover - third-party defensive path
        raw_toc = []
    nodes: list[dict[str, Any]] = []
    for index, item in enumerate(raw_toc, start=1):
        if not isinstance(item, (list, tuple)) or len(item) < 3:
            continue
        try:
            level = max(int(item[0]), 1)
            title = _sanitize_text(item[1])
            page_ordinal = max(int(item[2]), 1)
        except (TypeError, ValueError):
            continue
        if not title:
            continue
        nodes.append(
            {
                "node_id": f"outline-{index:03d}",
                "level": level,
                "title": title,
                "page_ordinal": page_ordinal,
            }
        )
    return nodes


def _suppress_repeated_heading_candidates(
    candidates_by_page: dict[int, list[dict[str, Any]]],
) -> dict[int, list[dict[str, Any]]]:
    counts = Counter(
        _slugify_text(candidate["text"])
        for candidates in candidates_by_page.values()
        for candidate in candidates
        if candidate.get("confidence") != "high"
    )
    filtered: dict[int, list[dict[str, Any]]] = {}
    for page_ordinal, candidates in candidates_by_page.items():
        filtered[page_ordinal] = []
        for candidate in candidates:
            slug = _slugify_text(candidate["text"])
            if counts.get(slug, 0) >= 3 and candidate.get("confidence") != "high":
                continue
            filtered[page_ordinal].append(candidate)
    return filtered


def _page_section_paths(
    *,
    page_count: int,
    outline_nodes: list[dict[str, Any]],
    heading_candidates_by_page: dict[int, list[dict[str, Any]]],
) -> dict[int, list[str]]:
    nodes_by_page: defaultdict[int, list[dict[str, Any]]] = defaultdict(list)
    for node in outline_nodes:
        page_ordinal = node.get("page_ordinal")
        if isinstance(page_ordinal, int):
            nodes_by_page[page_ordinal].append(node)
    active_stack: list[dict[str, Any]] = []
    section_paths: dict[int, list[str]] = {}
    for page_ordinal in range(1, page_count + 1):
        page_nodes = sorted(
            nodes_by_page.get(page_ordinal, []),
            key=lambda item: (int(item.get("level", 1)), str(item.get("title", ""))),
        )
        if page_nodes:
            for node in page_nodes:
                level = max(int(node.get("level", 1)), 1)
                active_stack = [item for item in active_stack if int(item.get("level", 1)) < level]
                active_stack.append({"level": level, "title": str(node["title"])})
        elif heading_candidates_by_page.get(page_ordinal):
            candidate = heading_candidates_by_page[page_ordinal][0]
            level = max(int(candidate.get("level", 1)), 1)
            title = str(candidate.get("text") or "")
            if title and (not active_stack or active_stack[-1].get("title") != title):
                active_stack = [item for item in active_stack if int(item.get("level", 1)) < level]
                active_stack.append({"level": level, "title": title})
        section_paths[page_ordinal] = [
            str(item["title"]) for item in active_stack if item.get("title")
        ]
    return section_paths


def _caption_text(text: str) -> str | None:
    compact = _compact_text(text, limit=160)
    if not compact:
        return None
    if not CAPTION_PREFIX_PATTERN.match(compact):
        return None
    if len(compact.split()) > 18:
        return None
    return compact


def _heading_aliases_from_path(section_path: list[str]) -> list[str]:
    aliases: list[str] = []
    for index in range(1, len(section_path) + 1):
        aliases.append(" / ".join(section_path[:index]))
    return deduplicate_strings(aliases)


def _header_key_from_text(text: str) -> str:
    lines = [_sanitize_text(line) for line in str(text).splitlines() if _sanitize_text(line)]
    if not lines:
        return ""
    return " | ".join(lines[:2]).lower()


def _procedure_spans_for_region(
    region: dict[str, Any],
    *,
    unit_id: str,
) -> list[dict[str, Any]]:
    text = str(region.get("linked_text") or "")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) < 2:
        return []
    ordered_steps = [ORDERED_STEP_PATTERN.match(line) for line in lines]
    if sum(match is not None for match in ordered_steps) >= 2:
        numbers = [int(match.group(1)) for match in ordered_steps if match is not None]
        return [
            {
                "unit_id": unit_id,
                "artifact_id": region.get("artifact_id"),
                "step_kind": "ordered",
                "start_number": min(numbers),
                "end_number": max(numbers),
                "step_count": len(numbers),
                "text_excerpt": _compact_text("\n".join(lines[:4]), limit=160),
                "confidence": "medium" if len(numbers) >= 3 else "low",
            }
        ]
    bullet_count = sum(1 for line in lines if BULLET_STEP_PATTERN.match(line))
    if bullet_count >= 3:
        return [
            {
                "unit_id": unit_id,
                "artifact_id": region.get("artifact_id"),
                "step_kind": "bullet",
                "start_number": None,
                "end_number": None,
                "step_count": bullet_count,
                "text_excerpt": _compact_text("\n".join(lines[:4]), limit=160),
                "confidence": "low",
            }
        ]
    return []


def _words_near_bbox(
    words: list[dict[str, Any]],
    bbox: dict[str, float],
    *,
    padding: float,
    limit: int = 24,
) -> str:
    nearby = [
        str(record["text"])
        for record in words
        if isinstance(record, dict)
        and isinstance(record.get("bbox"), dict)
        and _bboxes_overlap(record["bbox"], bbox, padding=padding)
        and isinstance(record.get("text"), str)
    ]
    return _compact_text(" ".join(nearby[:limit]), limit=220)


def _drawing_bbox(drawing: dict[str, Any]) -> dict[str, float] | None:
    bbox = _bbox_from_pdf_rect(drawing.get("rect"))
    if bbox is not None:
        return bbox
    item_boxes = [
        item_bbox
        for item in drawing.get("items", [])
        if isinstance(item, (list, tuple)) and len(item) >= 2
        for item_bbox in [_bbox_from_pdf_rect(item[1])]
        if item_bbox is not None
    ]
    return _union_bbox(item_boxes)


def _cluster_pdf_drawings(drawings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    clusters: list[dict[str, Any]] = []
    for drawing in drawings:
        if not isinstance(drawing, dict):
            continue
        bbox = _drawing_bbox(drawing)
        if bbox is None:
            continue
        merged = False
        for cluster in clusters:
            cluster_bbox = cluster.get("bbox")
            if not isinstance(cluster_bbox, dict):
                continue
            if _bboxes_overlap(cluster_bbox, bbox, padding=18.0):
                cluster["boxes"].append(bbox)
                cluster["drawings"].append(drawing)
                cluster["bbox"] = _union_bbox(cluster["boxes"])
                merged = True
                break
        if not merged:
            clusters.append({"bbox": bbox, "boxes": [bbox], "drawings": [drawing]})

    changed = True
    while changed:
        changed = False
        merged_clusters: list[dict[str, Any]] = []
        while clusters:
            cluster = clusters.pop(0)
            cluster_bbox = cluster.get("bbox")
            if not isinstance(cluster_bbox, dict):
                continue
            merged_index = next(
                (
                    index
                    for index, existing in enumerate(merged_clusters)
                    if isinstance(existing.get("bbox"), dict)
                    and _bboxes_overlap(existing["bbox"], cluster_bbox, padding=24.0)
                ),
                None,
            )
            if merged_index is None:
                merged_clusters.append(cluster)
                continue
            merged_clusters[merged_index]["boxes"].extend(cluster["boxes"])
            merged_clusters[merged_index]["drawings"].extend(cluster["drawings"])
            merged_clusters[merged_index]["bbox"] = _union_bbox(
                merged_clusters[merged_index]["boxes"]
            )
            changed = True
        clusters = merged_clusters
    return clusters


def _looks_chart_like(*, text: str, drawing_count: int) -> bool:
    lowered = text.lower()
    score = 0
    if any(token in lowered for token in CHART_TOKENS):
        score += 2
    if any(token in lowered for token in MONTH_TOKENS):
        score += 1
    if len(re.findall(r"\b\d+(?:\.\d+)?%?\b", lowered)) >= 2:
        score += 1
    if drawing_count >= 3:
        score += 1
    return score >= 3


def _salient_pdf_text_region(
    text: str,
    *,
    bbox: dict[str, float],
    page_width: float,
    page_height: float,
) -> bool:
    if len(text.split()) >= 4:
        return True
    if _bbox_area(bbox) >= (page_width * page_height * 0.025):
        return True
    return float(bbox["y0"]) <= (page_height * 0.18)


def compile_pdf_visual_artifacts(
    source_dir: Path,
    *,
    source_id: str,
    pdf_path: Path,
    units: list[dict[str, Any]],
    page_texts: list[str] | None = None,
) -> dict[str, Any]:
    visual_dir = source_dir / "visual_layout"
    visual_dir.mkdir(parents=True, exist_ok=True)
    pymupdf = _open_pymupdf()
    unit_updates: dict[str, dict[str, Any]] = {}
    layout_assets: list[str] = []
    pymupdf_warnings: list[str] = []
    artifact_entries: list[dict[str, Any]] = []
    outline_nodes: list[dict[str, Any]] = []
    page_payloads: list[dict[str, Any]] = []
    try:
        pymupdf.TOOLS.reset_mupdf_warnings()
    except Exception:
        pass
    document = pymupdf.open(str(pdf_path))
    try:
        outline_nodes = _outline_nodes_from_pdf(document)
        for unit in units:
            unit_id = str(unit.get("unit_id") or "")
            ordinal = unit.get("ordinal")
            if (
                not unit_id
                or not isinstance(ordinal, int)
                or ordinal <= 0
                or ordinal > len(document)
            ):
                continue
            page = document[ordinal - 1]
            page_width = float(page.rect.width)
            page_height = float(page.rect.height)
            page_area = max(page_width * page_height, 1.0)
            extracted_text = (
                str(page_texts[ordinal - 1])
                if isinstance(page_texts, list) and ordinal - 1 < len(page_texts)
                else ""
            )
            render_asset = unit.get("rendered_asset")
            render_assets = [render_asset] if isinstance(render_asset, str) and render_asset else []
            regions: list[dict[str, Any]] = []
            page_artifacts: list[dict[str, Any]] = []
            artifact_counter: Counter[str] = Counter()
            texts_for_role: list[str] = []
            picture_boxes: list[dict[str, float]] = []
            graphic_boxes: list[dict[str, float]] = []
            word_records = _pdf_word_records(page)
            word_lines = _pdf_word_lines(word_records)
            page_dict = page.get_text("dict")
            text_lines = _pdf_text_lines(page_dict)
            heading_candidates, _body_size = _pdf_heading_candidates(
                text_lines,
                page_height=page_height,
            )
            for block in page_dict.get("blocks", []):
                if not isinstance(block, dict):
                    continue
                block_type = int(block.get("type", -1))
                bbox = _bbox_from_pdf_rect(block.get("bbox"))
                if bbox is None:
                    continue
                if block_type == 0:
                    text = "\n".join(
                        _sanitize_text(span.get("text"))
                        for line in block.get("lines", [])
                        if isinstance(line, dict)
                        for span in line.get("spans", [])
                        if isinstance(span, dict) and _sanitize_text(span.get("text"))
                    ).strip()
                    if not text:
                        continue
                    texts_for_role.append(text)
                    artifact_counter["text-region"] += 1
                    artifact_id = stable_artifact_id(
                        unit_id,
                        "text-region",
                        artifact_counter["text-region"],
                    )
                    title = artifact_title_from_text(
                        text,
                        artifact_type="text-region",
                        fallback=f"Text Region {artifact_counter['text-region']}",
                    )
                    region = {
                        "artifact_id": artifact_id,
                        "artifact_type": "text-region",
                        "title": title,
                        "bbox": bbox,
                        "normalized_bbox": normalize_bbox(
                            bbox, width=page_width, height=page_height
                        ),
                        "linked_text": text,
                        "available_channels": ["text", "render", "structure"],
                        "render_assets": render_assets,
                        "derivation_mode": "deterministic",
                    }
                    regions.append(region)
                    if _salient_pdf_text_region(
                        text,
                        bbox=bbox,
                        page_width=page_width,
                        page_height=page_height,
                    ):
                        page_artifacts.append(
                            {
                                "artifact_id": artifact_id,
                                "artifact_type": "text-region",
                                "unit_id": unit_id,
                                "title": title,
                                "locator_aliases": artifact_locator_aliases(
                                    artifact_type="text-region",
                                    title=title,
                                    unit_title=str(unit.get("title") or ""),
                                ),
                                "available_channels": ["text", "render", "structure"],
                                "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                                "graph_promoted": False,
                                "bbox": bbox,
                                "normalized_bbox": normalize_bbox(
                                    bbox,
                                    width=page_width,
                                    height=page_height,
                                ),
                                "render_assets": render_assets,
                                "render_page_span": {"start": ordinal, "end": ordinal},
                                "linked_text": text,
                                "searchable_text": "\n".join([title, text]),
                                "derivation_mode": "deterministic",
                                "semantic_gap_hints": [],
                            }
                        )
                elif block_type == 1:
                    nearby_text = _words_near_bbox(word_records, bbox, padding=28.0)
                    inferred_type = (
                        "chart"
                        if _looks_chart_like(text=nearby_text, drawing_count=0)
                        else "picture"
                    )
                    artifact_counter[inferred_type] += 1
                    artifact_id = stable_artifact_id(
                        unit_id,
                        inferred_type,
                        artifact_counter[inferred_type],
                    )
                    title = artifact_title_from_text(
                        nearby_text,
                        artifact_type=inferred_type,
                        fallback=(
                            f"Chart Region {artifact_counter[inferred_type]}"
                            if inferred_type == "chart"
                            else f"Image Region {artifact_counter[inferred_type]}"
                        ),
                    )
                    available_channels = ["render", "media", "structure"]
                    if nearby_text or inferred_type == "chart":
                        available_channels = ["text", "render", "media", "structure"]
                    region = {
                        "artifact_id": artifact_id,
                        "artifact_type": inferred_type,
                        "title": title,
                        "bbox": bbox,
                        "normalized_bbox": normalize_bbox(
                            bbox, width=page_width, height=page_height
                        ),
                        "linked_text": nearby_text,
                        "available_channels": available_channels,
                        "render_assets": render_assets,
                        "nearby_text": nearby_text,
                        "visual_hints": ["chart-like"] if inferred_type == "chart" else [],
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": [],
                    }
                    regions.append(region)
                    if inferred_type == "picture":
                        picture_boxes.append(bbox)
                    graphic_boxes.append(bbox)
                    page_artifacts.append(
                        {
                            "artifact_id": artifact_id,
                            "artifact_type": inferred_type,
                            "unit_id": unit_id,
                            "title": title,
                            "locator_aliases": artifact_locator_aliases(
                                artifact_type=inferred_type,
                                title=title,
                                unit_title=str(unit.get("title") or ""),
                            ),
                            "available_channels": available_channels,
                            "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                            "graph_promoted": inferred_type == "chart",
                            "bbox": bbox,
                            "normalized_bbox": normalize_bbox(
                                bbox, width=page_width, height=page_height
                            ),
                            "render_assets": render_assets,
                            "render_page_span": {"start": ordinal, "end": ordinal},
                            "linked_text": nearby_text,
                            "visual_hints": ["chart-like"] if inferred_type == "chart" else [],
                            "searchable_text": "\n".join(
                                part for part in [title, nearby_text] if part
                            ),
                            "derivation_mode": "deterministic",
                            "semantic_gap_hints": [],
                        }
                    )
            try:
                table_finder = page.find_tables()
                tables = getattr(table_finder, "tables", [])
            except Exception:  # pragma: no cover - third-party defensive path
                tables = []
            table_bboxes: list[dict[str, float]] = []
            for index, table in enumerate(tables, start=1):
                bbox = _bbox_from_pdf_rect(getattr(table, "bbox", None))
                if bbox is None:
                    continue
                table_rows = table.extract() or []
                flattened = [
                    _sanitize_text(cell)
                    for row in table_rows
                    if isinstance(row, list)
                    for cell in row
                    if _sanitize_text(cell)
                ]
                table_text = "\n".join(flattened[:12]).strip()
                artifact_counter["table"] += 1
                artifact_id = stable_artifact_id(unit_id, "table", artifact_counter["table"])
                title = artifact_title_from_text(
                    table_text,
                    artifact_type="table",
                    fallback=f"Table {index}",
                )
                nearby_text = _words_near_bbox(word_records, bbox, padding=18.0)
                table_bboxes.append(bbox)
                region = {
                    "artifact_id": artifact_id,
                    "artifact_type": "table",
                    "title": title,
                    "bbox": bbox,
                    "normalized_bbox": normalize_bbox(bbox, width=page_width, height=page_height),
                    "linked_text": table_text or nearby_text,
                    "available_channels": ["text", "render", "structure"],
                    "render_assets": render_assets,
                    "table_preview": table_rows[:8],
                    "nearby_text": nearby_text,
                    "derivation_mode": "deterministic",
                    "semantic_gap_hints": [],
                }
                regions.append(region)
                page_artifacts.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": "table",
                        "unit_id": unit_id,
                        "title": title,
                        "locator_aliases": artifact_locator_aliases(
                            artifact_type="table",
                            title=title,
                            unit_title=str(unit.get("title") or ""),
                        ),
                        "available_channels": ["text", "render", "structure"],
                        "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                        "graph_promoted": artifact_graph_promoted(
                            artifact_type="table",
                            high_confidence=True,
                        ),
                        "bbox": bbox,
                        "normalized_bbox": normalize_bbox(
                            bbox, width=page_width, height=page_height
                        ),
                        "render_assets": render_assets,
                        "render_page_span": {"start": ordinal, "end": ordinal},
                        "linked_text": table_text or nearby_text,
                        "searchable_text": "\n".join(
                            part for part in [title, table_text, nearby_text] if part
                        ),
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": [],
                    }
                )
            for manual_table in _manual_pdf_table_candidates(
                word_lines,
                excluded_bboxes=table_bboxes,
            ):
                bbox = manual_table["bbox"]
                preview_rows = manual_table["rows"]
                flattened = [
                    _sanitize_text(cell)
                    for row in preview_rows
                    for cell in row
                    if _sanitize_text(cell)
                ]
                table_text = "\n".join(flattened[:12]).strip()
                nearby_text = _words_near_bbox(word_records, bbox, padding=18.0)
                artifact_counter["table"] += 1
                artifact_id = stable_artifact_id(unit_id, "table", artifact_counter["table"])
                title = artifact_title_from_text(
                    table_text,
                    artifact_type="table",
                    fallback=f"Table {artifact_counter['table']}",
                )
                table_bboxes.append(bbox)
                regions.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": "table",
                        "title": title,
                        "bbox": bbox,
                        "normalized_bbox": normalize_bbox(
                            bbox, width=page_width, height=page_height
                        ),
                        "linked_text": table_text or nearby_text,
                        "available_channels": ["text", "render", "structure"],
                        "render_assets": render_assets,
                        "table_preview": preview_rows[:8],
                        "nearby_text": nearby_text,
                        "visual_hints": ["borderless-table-like"],
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": [],
                    }
                )
                page_artifacts.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": "table",
                        "unit_id": unit_id,
                        "title": title,
                        "locator_aliases": artifact_locator_aliases(
                            artifact_type="table",
                            title=title,
                            unit_title=str(unit.get("title") or ""),
                        ),
                        "available_channels": ["text", "render", "structure"],
                        "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                        "graph_promoted": artifact_graph_promoted(
                            artifact_type="table",
                            high_confidence=True,
                        ),
                        "bbox": bbox,
                        "normalized_bbox": normalize_bbox(
                            bbox, width=page_width, height=page_height
                        ),
                        "render_assets": render_assets,
                        "render_page_span": {"start": ordinal, "end": ordinal},
                        "linked_text": table_text or nearby_text,
                        "visual_hints": ["borderless-table-like"],
                        "searchable_text": "\n".join(
                            part for part in [title, table_text, nearby_text] if part
                        ),
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": [],
                    }
                )
            try:
                drawings = page.get_drawings()
            except Exception:  # pragma: no cover - third-party defensive path
                drawings = []
            drawing_clusters = _cluster_pdf_drawings(drawings)
            for cluster in drawing_clusters:
                bbox = cluster.get("bbox")
                if not isinstance(bbox, dict):
                    continue
                if _bbox_area(bbox) < (page_width * page_height * 0.01):
                    continue
                if any(
                    _bboxes_overlap(bbox, table_bbox, padding=8.0) for table_bbox in table_bboxes
                ):
                    continue
                nearby_text = _words_near_bbox(word_records, bbox, padding=26.0)
                drawing_count = len(cluster.get("drawings", []))
                inferred_type = (
                    "chart"
                    if _looks_chart_like(text=nearby_text, drawing_count=drawing_count)
                    else "major-region"
                )
                artifact_counter[inferred_type] += 1
                artifact_id = stable_artifact_id(
                    unit_id,
                    inferred_type,
                    artifact_counter[inferred_type],
                )
                title = artifact_title_from_text(
                    nearby_text,
                    artifact_type=inferred_type,
                    fallback=(
                        f"Chart Cluster {artifact_counter[inferred_type]}"
                        if inferred_type == "chart"
                        else f"Graphic Cluster {artifact_counter[inferred_type]}"
                    ),
                )
                visual_hints = ["graphic-cluster"]
                if inferred_type == "chart":
                    visual_hints.insert(0, "chart-like")
                regions.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": inferred_type,
                        "title": title,
                        "bbox": bbox,
                        "normalized_bbox": normalize_bbox(
                            bbox, width=page_width, height=page_height
                        ),
                        "linked_text": nearby_text,
                        "available_channels": ["text", "render", "structure"],
                        "render_assets": render_assets,
                        "nearby_text": nearby_text,
                        "visual_hints": visual_hints,
                        "drawing_count": drawing_count,
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": [],
                    }
                )
                graphic_boxes.append(bbox)
                page_artifacts.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": inferred_type,
                        "unit_id": unit_id,
                        "title": title,
                        "locator_aliases": artifact_locator_aliases(
                            artifact_type=inferred_type,
                            title=title,
                            unit_title=str(unit.get("title") or ""),
                        ),
                        "available_channels": ["text", "render", "structure"],
                        "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                        "graph_promoted": artifact_graph_promoted(
                            artifact_type=inferred_type,
                            high_confidence=True,
                        ),
                        "bbox": bbox,
                        "normalized_bbox": normalize_bbox(
                            bbox, width=page_width, height=page_height
                        ),
                        "render_assets": render_assets,
                        "render_page_span": {"start": ordinal, "end": ordinal},
                        "linked_text": nearby_text,
                        "visual_hints": visual_hints,
                        "searchable_text": "\n".join(part for part in [title, nearby_text] if part),
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": [],
                    }
                )
            text_layer_quality, page_gap_hints = _pdf_text_layer_quality(
                extracted_text=extracted_text,
                word_records=word_records,
                picture_boxes=picture_boxes,
                graphic_boxes=graphic_boxes,
                page_area=page_area,
            )
            page_image_artifact_id: str | None = None
            if {"image-only-page", "scanned-page-like"} & set(page_gap_hints) or (
                "weak-text-layer" in page_gap_hints and picture_boxes
            ):
                artifact_counter["page-image"] += 1
                page_image_artifact_id = stable_artifact_id(
                    unit_id,
                    "page-image",
                    artifact_counter["page-image"],
                )
                full_page_bbox = {
                    "x0": 0.0,
                    "y0": 0.0,
                    "x1": float(page_width),
                    "y1": float(page_height),
                }
                page_image_title = f"{unit.get('title') or f'Page {ordinal}'} image"
                page_image_region = {
                    "artifact_id": page_image_artifact_id,
                    "artifact_type": "page-image",
                    "title": page_image_title,
                    "bbox": full_page_bbox,
                    "normalized_bbox": normalize_bbox(
                        full_page_bbox,
                        width=page_width,
                        height=page_height,
                    ),
                    "linked_text": "",
                    "available_channels": ["render", "media", "structure"],
                    "render_assets": render_assets,
                    "visual_hints": ["full-page-image"],
                    "derivation_mode": "deterministic",
                    "semantic_gap_hints": list(page_gap_hints),
                    "text_layer_quality": text_layer_quality,
                }
                regions.append(page_image_region)
                page_artifacts.append(
                    {
                        "artifact_id": page_image_artifact_id,
                        "artifact_type": "page-image",
                        "unit_id": unit_id,
                        "title": page_image_title,
                        "locator_aliases": artifact_locator_aliases(
                            artifact_type="page-image",
                            title=page_image_title,
                            unit_title=str(unit.get("title") or ""),
                        ),
                        "available_channels": ["render", "media", "structure"],
                        "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                        "graph_promoted": False,
                        "bbox": full_page_bbox,
                        "normalized_bbox": normalize_bbox(
                            full_page_bbox,
                            width=page_width,
                            height=page_height,
                        ),
                        "render_assets": render_assets,
                        "render_page_span": {"start": ordinal, "end": ordinal},
                        "linked_text": "",
                        "visual_hints": ["full-page-image"],
                        "searchable_text": "\n".join(
                            [
                                page_image_title,
                                " ".join(page_gap_hints),
                            ]
                        ).strip(),
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": list(page_gap_hints),
                        "text_layer_quality": text_layer_quality,
                    }
                )
            for artifact in page_artifacts:
                artifact["text_layer_quality"] = text_layer_quality
                artifact["semantic_gap_hints"] = _pdf_artifact_semantic_gap_hints(
                    artifact_type=str(artifact.get("artifact_type") or ""),
                    linked_text=str(artifact.get("linked_text") or ""),
                    caption_text=str(artifact.get("caption_text") or ""),
                    page_gap_hints=page_gap_hints,
                )
            for region in regions:
                region["text_layer_quality"] = text_layer_quality
                region["semantic_gap_hints"] = _pdf_artifact_semantic_gap_hints(
                    artifact_type=str(region.get("artifact_type") or ""),
                    linked_text=str(region.get("linked_text") or ""),
                    caption_text=str(region.get("caption_text") or ""),
                    page_gap_hints=page_gap_hints,
                )
            role_hints = _role_hints_from_text(
                texts_for_role
                + [
                    str(artifact.get("linked_text") or "")
                    for artifact in page_artifacts
                    if isinstance(artifact, dict)
                ],
                artifact_types=[str(artifact["artifact_type"]) for artifact in page_artifacts],
            )
            page_payloads.append(
                {
                    "unit_id": unit_id,
                    "ordinal": ordinal,
                    "unit_title": str(unit.get("title") or ""),
                    "render_assets": render_assets,
                    "page_size": {"width": page_width, "height": page_height},
                    "regions": regions,
                    "page_artifacts": page_artifacts,
                    "role_hints": role_hints,
                    "heading_candidates": heading_candidates,
                    "text_layer_quality": text_layer_quality,
                    "semantic_gap_hints": page_gap_hints,
                    "page_image_artifact_id": page_image_artifact_id,
                }
            )
    finally:
        document.close()
        pymupdf_warnings = _consume_pymupdf_warnings(pymupdf)

    heading_candidates_by_page = _suppress_repeated_heading_candidates(
        {
            int(page_payload["ordinal"]): list(page_payload["heading_candidates"])
            for page_payload in page_payloads
        }
    )
    if not outline_nodes:
        derived_nodes: list[dict[str, Any]] = []
        for page_payload in page_payloads:
            ordinal = int(page_payload["ordinal"])
            candidates = heading_candidates_by_page.get(ordinal, [])
            if not candidates:
                continue
            candidate = candidates[0]
            if candidate.get("confidence") not in {"high", "medium"}:
                continue
            derived_nodes.append(
                {
                    "node_id": f"outline-derived-{len(derived_nodes) + 1:03d}",
                    "level": int(candidate.get("level", 1)),
                    "title": str(candidate["text"]),
                    "page_ordinal": ordinal,
                }
            )
        outline_nodes = derived_nodes

    section_paths = _page_section_paths(
        page_count=len(page_payloads),
        outline_nodes=outline_nodes,
        heading_candidates_by_page=heading_candidates_by_page,
    )
    caption_links: list[dict[str, Any]] = []
    continuation_links: list[dict[str, Any]] = []
    procedure_spans: list[dict[str, Any]] = []
    document_role_hints: list[str] = []
    continuation_index = 0
    procedure_index = 0
    artifact_lookup: dict[str, dict[str, Any]] = {}
    page_contexts: list[dict[str, Any]] = []

    for page_payload in page_payloads:
        ordinal = int(page_payload["ordinal"])
        unit_id = str(page_payload["unit_id"])
        section_path = section_paths.get(ordinal, [])
        heading_aliases = _heading_aliases_from_path(section_path)
        unit_locator_aliases = deduplicate_strings(
            heading_aliases + ([section_path[-1]] if section_path else []) + [f"Page {ordinal}"]
        )
        unit_updates[unit_id] = {
            "render_assets": list(page_payload["render_assets"]),
            "render_page_span": {"start": ordinal, "end": ordinal},
            "section_path": section_path,
            "heading_aliases": heading_aliases,
            "locator_aliases": unit_locator_aliases,
            "semantic_gap_hints": list(page_payload.get("semantic_gap_hints", [])),
            "text_layer_quality": page_payload.get("text_layer_quality"),
        }
        document_role_hints.extend(
            hint for hint in page_payload["role_hints"] if isinstance(hint, str) and hint
        )
        artifact_lookup.update(
            {
                str(artifact["artifact_id"]): artifact
                for artifact in page_payload["page_artifacts"]
                if isinstance(artifact.get("artifact_id"), str)
            }
        )

        caption_artifact_ids: list[str] = []
        for artifact in page_payload["page_artifacts"]:
            artifact["section_path"] = list(section_path)
            artifact["caption_text"] = None
            artifact["continuation_group_ids"] = []
            artifact["procedure_hints"] = []
            artifact["derivation_mode"] = "deterministic"
            artifact["locator_aliases"] = deduplicate_strings(
                list(artifact.get("locator_aliases", [])) + heading_aliases
            )
            if section_path:
                searchable = str(artifact.get("searchable_text") or "")
                artifact["searchable_text"] = "\n".join(
                    part for part in [searchable, " / ".join(section_path)] if part
                )
        for region in page_payload["regions"]:
            region["section_path"] = list(section_path)
            region["caption_text"] = None
            region["continuation_group_ids"] = []
            region["procedure_hints"] = []
            region["derivation_mode"] = "deterministic"

        caption_candidates = [
            artifact
            for artifact in page_payload["page_artifacts"]
            if artifact.get("artifact_type") == "text-region"
            and isinstance(_caption_text(str(artifact.get("linked_text") or "")), str)
        ]
        target_candidates = [
            artifact
            for artifact in page_payload["page_artifacts"]
            if artifact.get("artifact_type") in {"table", "chart", "picture", "major-region"}
        ]
        for caption in caption_candidates:
            caption_bbox = caption.get("bbox")
            if not isinstance(caption_bbox, dict):
                continue
            caption_text = _caption_text(str(caption.get("linked_text") or ""))
            if not caption_text:
                continue
            best_target: dict[str, Any] | None = None
            best_distance: float | None = None
            for target in target_candidates:
                target_bbox = target.get("bbox")
                if not isinstance(target_bbox, dict):
                    continue
                horizontal_overlap = min(
                    float(caption_bbox["x1"]),
                    float(target_bbox["x1"]),
                ) - max(float(caption_bbox["x0"]), float(target_bbox["x0"]))
                if horizontal_overlap <= 0:
                    continue
                vertical_gap = min(
                    abs(float(target_bbox["y0"]) - float(caption_bbox["y1"])),
                    abs(float(caption_bbox["y0"]) - float(target_bbox["y1"])),
                )
                if vertical_gap > 88.0:
                    continue
                if best_distance is None or vertical_gap < best_distance:
                    best_distance = vertical_gap
                    best_target = target
            if best_target is None:
                continue
            best_target["caption_text"] = caption_text
            best_target["searchable_text"] = "\n".join(
                part
                for part in [str(best_target.get("searchable_text") or ""), caption_text]
                if part
            )
            best_target["locator_aliases"] = deduplicate_strings(
                list(best_target.get("locator_aliases", [])) + [caption_text]
            )
            for region in page_payload["regions"]:
                if region.get("artifact_id") == best_target.get("artifact_id"):
                    region["caption_text"] = caption_text
                    break
            caption_artifact_ids.append(str(best_target["artifact_id"]))
            caption_links.append(
                {
                    "caption_text": caption_text,
                    "caption_unit_id": unit_id,
                    "caption_artifact_id": caption.get("artifact_id"),
                    "target_unit_id": unit_id,
                    "target_artifact_id": best_target.get("artifact_id"),
                    "confidence": "high"
                    if best_distance is not None and best_distance <= 36.0
                    else "medium",
                }
            )

        page_procedure_spans: list[dict[str, Any]] = []
        for region in page_payload["regions"]:
            if region.get("artifact_type") != "text-region":
                continue
            for span in _procedure_spans_for_region(region, unit_id=unit_id):
                procedure_index += 1
                span["span_id"] = f"procedure-{procedure_index:03d}"
                page_procedure_spans.append(span)
                procedure_spans.append(span)
                if isinstance(region.get("artifact_id"), str):
                    region["procedure_hints"].append(
                        f"{span['step_kind']}-steps:{span.get('step_count', 0)}"
                    )
        for artifact in page_payload["page_artifacts"]:
            artifact_id = artifact.get("artifact_id")
            if not isinstance(artifact_id, str):
                continue
            for span in page_procedure_spans:
                if span.get("artifact_id") == artifact_id:
                    artifact["procedure_hints"].append(
                        f"{span['step_kind']}-steps:{span.get('step_count', 0)}"
                    )
                    searchable = str(artifact.get("searchable_text") or "")
                    artifact["searchable_text"] = "\n".join(
                        part for part in [searchable, str(span.get("text_excerpt") or "")] if part
                    )
            artifact["semantic_gap_hints"] = _pdf_artifact_semantic_gap_hints(
                artifact_type=str(artifact.get("artifact_type") or ""),
                linked_text=str(artifact.get("linked_text") or ""),
                caption_text=str(artifact.get("caption_text") or ""),
                page_gap_hints=list(page_payload.get("semantic_gap_hints", [])),
            )
        for region in page_payload["regions"]:
            region["semantic_gap_hints"] = _pdf_artifact_semantic_gap_hints(
                artifact_type=str(region.get("artifact_type") or ""),
                linked_text=str(region.get("linked_text") or ""),
                caption_text=str(region.get("caption_text") or ""),
                page_gap_hints=list(page_payload.get("semantic_gap_hints", [])),
            )
        page_contexts.append(
            {
                "unit_id": unit_id,
                "page_ordinal": ordinal,
                "section_path": list(section_path),
                "heading_candidates": list(heading_candidates_by_page.get(ordinal, [])),
                "caption_artifact_ids": deduplicate_strings(caption_artifact_ids),
                "continuation_group_ids": [],
                "procedure_spans": [dict(span) for span in page_procedure_spans],
                "role_hints": list(page_payload["role_hints"]),
                "text_layer_quality": page_payload.get("text_layer_quality"),
                "semantic_gap_hints": list(page_payload.get("semantic_gap_hints", [])),
                "page_image_artifact_id": page_payload.get("page_image_artifact_id"),
            }
        )

    context_lookup = {
        str(context["unit_id"]): context
        for context in page_contexts
        if isinstance(context.get("unit_id"), str)
    }
    for index in range(1, len(page_payloads)):
        previous_page = page_payloads[index - 1]
        current_page = page_payloads[index]
        previous_section_path = section_paths.get(int(previous_page["ordinal"]), [])
        current_section_path = section_paths.get(int(current_page["ordinal"]), [])
        previous_tables = [
            artifact
            for artifact in previous_page["page_artifacts"]
            if artifact.get("artifact_type") in {"table", "picture", "major-region"}
        ]
        current_tables = [
            artifact
            for artifact in current_page["page_artifacts"]
            if artifact.get("artifact_type") in {"table", "picture", "major-region"}
        ]
        for previous_artifact in previous_tables:
            previous_key = _slugify_text(
                previous_artifact.get("caption_text")
                or _header_key_from_text(previous_artifact.get("linked_text"))
            )
            if not previous_key:
                continue
            for current_artifact in current_tables:
                if previous_artifact.get("artifact_type") != current_artifact.get("artifact_type"):
                    continue
                current_key = _slugify_text(
                    current_artifact.get("caption_text")
                    or _header_key_from_text(current_artifact.get("linked_text"))
                )
                if not current_key or current_key != previous_key:
                    continue
                if (
                    previous_section_path
                    and current_section_path
                    and previous_section_path != current_section_path
                ):
                    continue
                continuation_index += 1
                group_id = (
                    f"{previous_artifact['artifact_type']}-continuation-{continuation_index:03d}"
                )
                previous_artifact["continuation_group_ids"].append(group_id)
                current_artifact["continuation_group_ids"].append(group_id)
                for region in previous_page["regions"]:
                    if region.get("artifact_id") == previous_artifact.get("artifact_id"):
                        region["continuation_group_ids"].append(group_id)
                for region in current_page["regions"]:
                    if region.get("artifact_id") == current_artifact.get("artifact_id"):
                        region["continuation_group_ids"].append(group_id)
                continuation_links.append(
                    {
                        "group_id": group_id,
                        "artifact_type": previous_artifact.get("artifact_type"),
                        "from_unit_id": previous_page["unit_id"],
                        "from_artifact_id": previous_artifact.get("artifact_id"),
                        "to_unit_id": current_page["unit_id"],
                        "to_artifact_id": current_artifact.get("artifact_id"),
                        "confidence": "high" if previous_artifact.get("caption_text") else "medium",
                        "reasons": [
                            reason
                            for reason in (
                                "caption-match" if previous_artifact.get("caption_text") else None,
                                "header-repeat"
                                if _header_key_from_text(previous_artifact.get("linked_text"))
                                else None,
                                "section-continuity"
                                if previous_section_path == current_section_path
                                else None,
                            )
                            if reason
                        ],
                    }
                )
                context_lookup[str(previous_page["unit_id"])]["continuation_group_ids"].append(
                    group_id
                )
                context_lookup[str(current_page["unit_id"])]["continuation_group_ids"].append(
                    group_id
                )
                break

        previous_spans = [
            span for span in procedure_spans if span.get("unit_id") == previous_page.get("unit_id")
        ]
        current_spans = [
            span for span in procedure_spans if span.get("unit_id") == current_page.get("unit_id")
        ]
        for previous_span in previous_spans:
            for current_span in current_spans:
                if (
                    previous_span.get("step_kind") != "ordered"
                    or current_span.get("step_kind") != "ordered"
                ):
                    continue
                if (
                    previous_section_path
                    and current_section_path
                    and previous_section_path != current_section_path
                ):
                    continue
                previous_end = previous_span.get("end_number")
                current_start = current_span.get("start_number")
                if not isinstance(previous_end, int) or not isinstance(current_start, int):
                    continue
                if current_start != previous_end + 1:
                    continue
                continuation_index += 1
                group_id = f"procedure-continuation-{continuation_index:03d}"
                previous_span["continuation_group_id"] = group_id
                current_span["continuation_group_id"] = group_id
                context_lookup[str(previous_page["unit_id"])]["continuation_group_ids"].append(
                    group_id
                )
                context_lookup[str(current_page["unit_id"])]["continuation_group_ids"].append(
                    group_id
                )
                continuation_links.append(
                    {
                        "group_id": group_id,
                        "artifact_type": "procedure",
                        "from_unit_id": previous_page["unit_id"],
                        "from_artifact_id": previous_span.get("artifact_id"),
                        "to_unit_id": current_page["unit_id"],
                        "to_artifact_id": current_span.get("artifact_id"),
                        "confidence": "medium",
                        "reasons": ["ordered-step-continuity"],
                    }
                )
                for artifact in (
                    previous_artifact for previous_artifact in previous_page["page_artifacts"]
                ):
                    if artifact.get("artifact_id") == previous_span.get("artifact_id"):
                        artifact["continuation_group_ids"].append(group_id)
                        artifact["procedure_hints"].append("continued-procedure")
                for artifact in (
                    current_artifact for current_artifact in current_page["page_artifacts"]
                ):
                    if artifact.get("artifact_id") == current_span.get("artifact_id"):
                        artifact["continuation_group_ids"].append(group_id)
                        artifact["procedure_hints"].append("continued-procedure")
                break

    for page_payload in page_payloads:
        unit_id = str(page_payload["unit_id"])
        layout_path = visual_dir / f"{unit_id}.json"
        layout_assets.append(str(layout_path.relative_to(source_dir)))
        write_json(
            layout_path,
            {
                "artifact_type": "visual-layout",
                "source_id": source_id,
                "unit_id": unit_id,
                "render_page_span": {
                    "start": int(page_payload["ordinal"]),
                    "end": int(page_payload["ordinal"]),
                },
                "render_assets": list(page_payload["render_assets"]),
                "page_size": dict(page_payload["page_size"]),
                "role_hints": list(page_payload["role_hints"]),
                "semantic_gap_hints": list(page_payload.get("semantic_gap_hints", [])),
                "regions": page_payload["regions"],
            },
        )
        artifact_entries.extend(page_payload["page_artifacts"])

    for context in page_contexts:
        context["continuation_group_ids"] = deduplicate_strings(
            [
                value
                for value in context.get("continuation_group_ids", [])
                if isinstance(value, str) and value
            ]
        )

    pdf_document = {
        "artifact_type": "pdf-document",
        "schema_version": 1,
        "generated_at": _utc_now(),
        "source_id": source_id,
        "derivation_mode": "deterministic",
        "outline_nodes": outline_nodes,
        "page_contexts": page_contexts,
        "caption_links": caption_links,
        "continuation_links": continuation_links,
        "procedure_spans": procedure_spans,
        "document_role_hints": deduplicate_strings(document_role_hints),
    }
    write_json(source_dir / "pdf_document.json", pdf_document)
    artifact_index = _write_artifact_index(
        source_dir, source_id=source_id, artifacts=artifact_entries
    )
    return {
        "artifact_index": artifact_index,
        "visual_layout_assets": layout_assets,
        "unit_updates": unit_updates,
        "pdf_document_asset": "pdf_document.json",
        "warnings": pymupdf_warnings,
    }


def _shape_bbox(shape: Any) -> dict[str, float] | None:
    try:
        return {
            "x0": float(shape.left),
            "y0": float(shape.top),
            "x1": float(shape.left + shape.width),
            "y1": float(shape.top + shape.height),
        }
    except Exception:
        return None


def _pptx_shape_type(shape: Any) -> str:
    if getattr(shape, "has_chart", False):
        return "chart"
    if getattr(shape, "has_table", False):
        return "table"
    shape_type = str(getattr(shape, "shape_type", "")).upper()
    if "PICTURE" in shape_type:
        return "picture"
    if "GROUP" in shape_type:
        return "group"
    if "CONNECTOR" in shape_type or (
        "LINE" in shape_type and hasattr(shape, "begin_x") and hasattr(shape, "end_x")
    ):
        return "connector"
    if "AUTO_SHAPE" in shape_type:
        return "auto-shape"
    if getattr(shape, "has_text_frame", False):
        return "text-box"
    return "other"


def _pptx_shape_text(shape: Any) -> str:
    if getattr(shape, "has_chart", False):
        chart = shape.chart
        parts: list[str] = []
        chart_title = getattr(chart, "chart_title", None)
        if getattr(chart, "has_title", False) and chart_title is not None:
            try:
                parts.append(_sanitize_text(chart_title.text_frame.text))
            except Exception:
                pass
        for series in _safe_pptx_chart_series(chart):
            name = _sanitize_text(getattr(series, "name", ""))
            if name:
                parts.append(name)
        return "\n".join(part for part in parts if part)
    if getattr(shape, "has_table", False):
        values = [
            _sanitize_text(cell.text)
            for row in shape.table.rows
            for cell in row.cells
            if _sanitize_text(cell.text)
        ]
        return "\n".join(values)
    if getattr(shape, "has_text_frame", False):
        return _sanitize_text(getattr(shape, "text", ""))
    return ""


def _pptx_chart_hints(shape: Any) -> dict[str, Any]:
    chart = shape.chart
    chart_type = _safe_pptx_chart_type(chart)
    series_names = [
        _sanitize_text(getattr(series, "name", ""))
        for series in _safe_pptx_chart_series(chart)
        if _sanitize_text(getattr(series, "name", ""))
    ]
    title = ""
    if getattr(chart, "has_title", False):
        try:
            title = _sanitize_text(chart.chart_title.text_frame.text)
        except Exception:
            title = ""
    return {
        "chart_type": chart_type or "unknown",
        "chart_title": title,
        "series_names": series_names[:8],
    }


def _safe_pptx_chart_series(chart: Any) -> list[Any]:
    """Return chart series without letting malformed chart XML abort publication."""
    try:
        series = getattr(chart, "series", [])
    except Exception:
        return []
    try:
        return list(series)
    except Exception:
        return []


def _safe_pptx_chart_type(chart: Any) -> str:
    """Return a stable chart-type hint for supported and unsupported python-pptx plots."""
    try:
        chart_type = _sanitize_text(str(getattr(chart, "chart_type", "")))
        if chart_type:
            return chart_type
    except Exception:
        pass
    try:
        plot_area = chart._chartSpace.plotArea
        for child in plot_area.iterchildren():
            tag = str(getattr(child, "tag", ""))
            if "}" in tag:
                tag = tag.rsplit("}", 1)[-1]
            if tag:
                return _sanitize_text(tag)
    except Exception:
        pass
    return "unknown"


def _pptx_table_preview(shape: Any) -> list[list[str]]:
    preview: list[list[str]] = []
    for row in list(shape.table.rows)[:6]:
        preview.append([_sanitize_text(cell.text) for cell in list(row.cells)[:8]])
    return preview


def _iter_pptx_shapes(shape_collection: Any) -> list[Any]:
    flattened: list[Any] = []
    for shape in shape_collection:
        flattened.append(shape)
        if _pptx_shape_type(shape) == "group":
            flattened.extend(_iter_pptx_shapes(shape.shapes))
    return flattened


def _pptx_endpoint_bbox(x: float, y: float, *, padding: float = 180000.0) -> dict[str, float]:
    return {"x0": x - padding, "y0": y - padding, "x1": x + padding, "y1": y + padding}


def _pptx_connector_links(
    connector_bbox: dict[str, float],
    *,
    begin_x: float | None,
    begin_y: float | None,
    end_x: float | None,
    end_y: float | None,
    targets: list[dict[str, Any]],
) -> list[str]:
    endpoint_boxes = []
    endpoint_points: list[tuple[float, float]] = []
    if isinstance(begin_x, (int, float)) and isinstance(begin_y, (int, float)):
        endpoint_boxes.append(_pptx_endpoint_bbox(float(begin_x), float(begin_y)))
        endpoint_points.append((float(begin_x), float(begin_y)))
    if isinstance(end_x, (int, float)) and isinstance(end_y, (int, float)):
        endpoint_boxes.append(_pptx_endpoint_bbox(float(end_x), float(end_y)))
        endpoint_points.append((float(end_x), float(end_y)))
    if not endpoint_boxes:
        endpoint_boxes = [connector_bbox]
    linked: list[str] = []
    for target in targets:
        bbox = target.get("bbox")
        artifact_id = target.get("artifact_id")
        if not isinstance(bbox, dict) or not isinstance(artifact_id, str):
            continue
        if any(
            _bboxes_overlap(endpoint_bbox, bbox, padding=0.0) for endpoint_bbox in endpoint_boxes
        ):
            linked.append(artifact_id)
            continue
        if endpoint_points:
            min_distance = min(
                min(
                    abs(point_x - float(bbox["x0"])),
                    abs(point_x - float(bbox["x1"])),
                )
                + min(
                    abs(point_y - float(bbox["y0"])),
                    abs(point_y - float(bbox["y1"])),
                )
                for point_x, point_y in endpoint_points
            )
            if min_distance <= 320000.0:
                linked.append(artifact_id)
    return deduplicate_strings(linked)


def _pptx_nearby_label(
    artifact: dict[str, Any],
    *,
    text_artifacts: list[dict[str, Any]],
) -> str | None:
    bbox = artifact.get("bbox")
    if not isinstance(bbox, dict):
        return None
    best_label: str | None = None
    best_distance: float | None = None
    for candidate in text_artifacts:
        candidate_id = candidate.get("artifact_id")
        if candidate_id == artifact.get("artifact_id"):
            continue
        candidate_bbox = candidate.get("bbox")
        candidate_text = _sanitize_text(candidate.get("linked_text"))
        if not isinstance(candidate_bbox, dict) or not candidate_text:
            continue
        if len(candidate_text.split()) > 12:
            continue
        horizontal_overlap = min(float(bbox["x1"]), float(candidate_bbox["x1"])) - max(
            float(bbox["x0"]),
            float(candidate_bbox["x0"]),
        )
        vertical_gap = min(
            abs(float(candidate_bbox["y0"]) - float(bbox["y1"])),
            abs(float(bbox["y0"]) - float(candidate_bbox["y1"])),
        )
        if horizontal_overlap <= 0 and vertical_gap > 180000.0:
            continue
        distance = (
            vertical_gap
            if horizontal_overlap > 0
            else abs(float(candidate_bbox["x0"]) - float(bbox["x1"]))
        )
        if distance > 320000.0:
            continue
        if best_distance is None or distance < best_distance:
            best_distance = distance
            best_label = candidate_text
    return best_label


def _pptx_group_child_order(shape: Any) -> list[str]:
    if _pptx_shape_type(shape) != "group":
        return []
    children = list(shape.shapes)
    ordered = sorted(
        children,
        key=lambda item: (
            float(getattr(item, "top", 0)),
            float(getattr(item, "left", 0)),
        ),
    )
    return [
        _sanitize_text(getattr(child, "name", "")) or _pptx_shape_type(child) for child in ordered
    ]


def compile_pptx_visual_artifacts(
    source_dir: Path,
    *,
    source_id: str,
    presentation: Any,
    units: list[dict[str, Any]],
) -> dict[str, Any]:
    visual_dir = source_dir / "visual_layout"
    visual_dir.mkdir(parents=True, exist_ok=True)
    artifact_entries: list[dict[str, Any]] = []
    unit_updates: dict[str, dict[str, Any]] = {}
    layout_assets: list[str] = []
    slide_width = float(getattr(presentation, "slide_width", 1) or 1)
    slide_height = float(getattr(presentation, "slide_height", 1) or 1)
    unit_lookup = {
        int(unit["ordinal"]): unit for unit in units if isinstance(unit.get("ordinal"), int)
    }
    for ordinal, slide in enumerate(list(presentation.slides), start=1):
        unit = unit_lookup.get(ordinal)
        if not isinstance(unit, dict):
            continue
        unit_id = str(unit.get("unit_id") or "")
        if not unit_id:
            continue
        render_asset = unit.get("rendered_asset")
        render_assets = [render_asset] if isinstance(render_asset, str) and render_asset else []
        render_page_span = unit.get("render_page_span")
        page_span: dict[str, int] | None
        if (
            isinstance(render_page_span, dict)
            and isinstance(render_page_span.get("start"), int)
            and isinstance(render_page_span.get("end"), int)
        ):
            page_span = {
                "start": int(render_page_span["start"]),
                "end": int(render_page_span["end"]),
            }
        else:
            render_ordinal = unit.get("render_ordinal")
            page_span = (
                {"start": render_ordinal, "end": render_ordinal}
                if isinstance(render_ordinal, int)
                else None
            )
        artifact_counter: Counter[str] = Counter()
        slide_artifacts: list[dict[str, Any]] = []
        regions: list[dict[str, Any]] = []
        texts_for_role: list[str] = []
        for shape in _iter_pptx_shapes(slide.shapes):
            artifact_type = _pptx_shape_type(shape)
            bbox = _shape_bbox(shape)
            if bbox is None:
                continue
            linked_text = _pptx_shape_text(shape)
            if linked_text:
                texts_for_role.append(linked_text)
            artifact_counter[artifact_type] += 1
            artifact_id = stable_artifact_id(
                unit_id, artifact_type, artifact_counter[artifact_type]
            )
            title = artifact_title_from_text(
                linked_text or getattr(shape, "name", ""),
                artifact_type=artifact_type,
                fallback=f"{artifact_type.title()} {artifact_counter[artifact_type]}",
            )
            region_payload: dict[str, Any] = {
                "artifact_id": artifact_id,
                "artifact_type": artifact_type,
                "title": title,
                "bbox": bbox,
                "normalized_bbox": normalize_bbox(bbox, width=slide_width, height=slide_height),
                "linked_text": linked_text,
                "available_channels": ["render", "structure"],
                "render_assets": render_assets,
                "derivation_mode": "deterministic",
                "semantic_gap_hints": [],
            }
            if artifact_type in {"text-box", "table", "chart"}:
                region_payload["available_channels"] = ["text", "render", "structure"]
            if artifact_type == "picture":
                region_payload["available_channels"] = ["render", "media", "structure"]
            if artifact_type == "chart":
                region_payload.update(_pptx_chart_hints(shape))
            if artifact_type == "table":
                region_payload["table_preview"] = _pptx_table_preview(shape)
            if artifact_type == "group":
                region_payload["group_child_order"] = _pptx_group_child_order(shape)
            if artifact_type == "connector":
                region_payload["begin_x"] = getattr(shape, "begin_x", None)
                region_payload["begin_y"] = getattr(shape, "begin_y", None)
                region_payload["end_x"] = getattr(shape, "end_x", None)
                region_payload["end_y"] = getattr(shape, "end_y", None)
            regions.append(region_payload)
            graph_promoted = artifact_graph_promoted(
                artifact_type=("major-region" if artifact_type == "group" else artifact_type),
                high_confidence=artifact_type in {"table", "chart", "group"},
            )
            search_parts = [title, linked_text]
            if artifact_type == "chart":
                search_parts.extend(region_payload.get("series_names", []))
                if isinstance(region_payload.get("chart_title"), str):
                    search_parts.append(region_payload["chart_title"])
            slide_artifacts.append(
                {
                    "artifact_id": artifact_id,
                    "artifact_type": artifact_type,
                    "unit_id": unit_id,
                    "title": title,
                    "locator_aliases": artifact_locator_aliases(
                        artifact_type=artifact_type,
                        title=title,
                        unit_title=str(unit.get("title") or ""),
                    ),
                    "available_channels": region_payload["available_channels"],
                    "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                    "graph_promoted": graph_promoted,
                    "bbox": bbox,
                    "normalized_bbox": normalize_bbox(bbox, width=slide_width, height=slide_height),
                    "render_assets": render_assets,
                    "render_page_span": page_span,
                    "linked_text": linked_text,
                    "searchable_text": "\n".join(
                        part for part in search_parts if isinstance(part, str) and part
                    ),
                    "visual_hints": [
                        hint
                        for hint in (
                            [region_payload["chart_type"]] if artifact_type == "chart" else []
                        )
                        if isinstance(hint, str) and hint
                    ],
                    "derivation_mode": "deterministic",
                    "semantic_gap_hints": [],
                }
            )
        region_lookup = {
            str(region["artifact_id"]): region
            for region in regions
            if isinstance(region.get("artifact_id"), str)
        }
        shape_targets = [
            artifact
            for artifact in slide_artifacts
            if artifact.get("artifact_type") not in {"connector", "text-box"}
        ]
        text_artifacts = [
            artifact for artifact in slide_artifacts if artifact.get("artifact_type") == "text-box"
        ]
        for artifact in slide_artifacts:
            artifact_type = str(artifact.get("artifact_type") or "")
            artifact_id_value = artifact.get("artifact_id")
            if not isinstance(artifact_id_value, str):
                continue
            artifact_id = artifact_id_value
            region = region_lookup.get(artifact_id)
            if artifact_type == "connector" and isinstance(region, dict):
                connected_shape_ids = _pptx_connector_links(
                    artifact["bbox"],
                    begin_x=region.get("begin_x"),
                    begin_y=region.get("begin_y"),
                    end_x=region.get("end_x"),
                    end_y=region.get("end_y"),
                    targets=shape_targets,
                )
                if connected_shape_ids:
                    artifact["linked_shape_artifact_ids"] = connected_shape_ids
                    artifact["searchable_text"] = "\n".join(
                        part
                        for part in [
                            artifact.get("searchable_text", ""),
                            " ".join(connected_shape_ids),
                        ]
                        if part
                    )
                    region["linked_shape_artifact_ids"] = connected_shape_ids
            if artifact_type in {"chart", "picture", "table", "auto-shape", "group"}:
                caption_text = _pptx_nearby_label(artifact, text_artifacts=text_artifacts)
                if caption_text:
                    artifact["caption_text"] = caption_text
                    artifact["locator_aliases"] = deduplicate_strings(
                        list(artifact.get("locator_aliases", [])) + [caption_text]
                    )
                    artifact["searchable_text"] = "\n".join(
                        part for part in [artifact.get("searchable_text", ""), caption_text] if part
                    )
                    if isinstance(region, dict):
                        region["caption_text"] = caption_text
        unit_gap_hints: list[str] = []
        picture_count = sum(
            1 for artifact in slide_artifacts if artifact.get("artifact_type") == "picture"
        )
        linked_text_length = len(
            " ".join(
                str(artifact.get("linked_text") or "")
                for artifact in slide_artifacts
                if isinstance(artifact.get("linked_text"), str)
            ).strip()
        )
        if picture_count >= 2:
            unit_gap_hints.append("picture-heavy-slide")
        if (
            any(
                artifact.get("artifact_type") in {"connector", "group"}
                for artifact in slide_artifacts
            )
            and linked_text_length <= 120
        ):
            unit_gap_hints.append("weak-label-slide")
        for artifact in slide_artifacts:
            artifact_type = str(artifact.get("artifact_type") or "")
            semantic_gap_hints: list[str] = []
            if artifact_type in {"connector", "group"} and linked_text_length <= 120:
                semantic_gap_hints.append("weak-label-slide")
            if artifact_type == "picture" and not artifact.get("caption_text"):
                semantic_gap_hints.append("rendered-only-picture")
            artifact["semantic_gap_hints"] = semantic_gap_hints
            region = region_lookup.get(str(artifact.get("artifact_id") or ""))
            if isinstance(region, dict):
                region["semantic_gap_hints"] = semantic_gap_hints
        role_hints = _role_hints_from_text(
            texts_for_role,
            artifact_types=[artifact["artifact_type"] for artifact in slide_artifacts],
        )
        layout_path = visual_dir / f"{unit_id}.json"
        layout_assets.append(str(layout_path.relative_to(source_dir)))
        write_json(
            layout_path,
            {
                "artifact_type": "visual-layout",
                "source_id": source_id,
                "unit_id": unit_id,
                "render_page_span": page_span,
                "render_assets": render_assets,
                "page_size": {"width": slide_width, "height": slide_height},
                "role_hints": role_hints,
                "semantic_gap_hints": deduplicate_strings(unit_gap_hints),
                "regions": regions,
            },
        )
        artifact_entries.extend(slide_artifacts)
        unit_updates[unit_id] = {
            "render_assets": render_assets,
            "render_page_span": page_span,
            "semantic_gap_hints": deduplicate_strings(unit_gap_hints),
        }
    artifact_index = _write_artifact_index(
        source_dir, source_id=source_id, artifacts=artifact_entries
    )
    return {
        "artifact_index": artifact_index,
        "visual_layout_assets": layout_assets,
        "unit_updates": unit_updates,
    }


def _defined_name_registry(workbook: Any) -> list[dict[str, Any]]:
    registry: list[dict[str, Any]] = []
    defined_names = getattr(workbook, "defined_names", {})
    keys = list(defined_names.keys()) if hasattr(defined_names, "keys") else []
    for name in keys:
        try:
            destinations = list(defined_names[name].destinations)
        except Exception:
            destinations = []
        registry.append(
            {
                "name": str(name),
                "destinations": [
                    {"sheet_name": str(sheet_name), "reference": str(reference)}
                    for sheet_name, reference in destinations
                ],
            }
        )
    return registry


def _iter_formula_cells(worksheet_formula: Any) -> list[dict[str, Any]]:
    formulas: list[dict[str, Any]] = []
    for row in worksheet_formula.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                refs = re.findall(r"(?:'([^']+)'|([A-Za-z0-9_ ]+))?!?([A-Z]{1,3}\d+)", value)
                formulas.append(
                    {
                        "cell": cell.coordinate,
                        "formula": value,
                        "references": [
                            {
                                "sheet_name": (
                                    sheet_a or sheet_b or worksheet_formula.title
                                ).strip(),
                                "cell": ref_cell,
                            }
                            for sheet_a, sheet_b, ref_cell in refs
                        ],
                    }
                )
    return formulas


def _infer_primary_region(worksheet_value: Any) -> str:
    if getattr(worksheet_value.auto_filter, "ref", None):
        return str(worksheet_value.auto_filter.ref)
    try:
        return str(worksheet_value.calculate_dimension())
    except Exception:
        max_column = max(worksheet_value.max_column, 1)
        max_row = max(worksheet_value.max_row, 1)
        return f"A1:{get_column_letter(max_column)}{max_row}"


def _tabular_regions(worksheet_value: Any) -> list[dict[str, Any]]:
    regions: list[dict[str, Any]] = []
    for name in worksheet_value.tables.keys():
        table = worksheet_value.tables[name]
        table_ref = table if isinstance(table, str) else getattr(table, "ref", None)
        if not isinstance(table_ref, str) or not table_ref:
            continue
        regions.append({"kind": "named-table", "name": name, "ref": table_ref})
    if not regions:
        regions.append(
            {
                "kind": "primary-range",
                "name": worksheet_value.title,
                "ref": _infer_primary_region(worksheet_value),
            }
        )
    return regions


def _header_names_from_region(worksheet_value: Any, region_ref: str) -> list[str]:
    min_col, min_row, max_col, _max_row = range_boundaries(region_ref)
    headers: list[str] = []
    for column in range(min_col, max_col + 1):
        header = _sanitize_text(worksheet_value.cell(row=min_row, column=column).value)
        headers.append(header or f"Column {get_column_letter(column)}")
    return headers


def _column_profiles(
    worksheet_value: Any,
    worksheet_formula: Any,
    region_ref: str,
) -> tuple[list[dict[str, Any]], list[str], list[str], list[str], list[dict[str, Any]]]:
    min_col, min_row, max_col, max_row = range_boundaries(region_ref)
    header_names = _header_names_from_region(worksheet_value, region_ref)
    metric_candidates: list[str] = []
    dimension_candidates: list[str] = []
    time_axis_candidates: list[str] = []
    profiles: list[dict[str, Any]] = []
    anomalies: list[dict[str, Any]] = []
    for offset, column in enumerate(range(min_col, max_col + 1), start=0):
        header = header_names[offset]
        samples = [
            worksheet_value.cell(row=row, column=column).value
            for row in range(min_row + 1, min(max_row, min_row + 25) + 1)
            if worksheet_value.cell(row=row, column=column).value not in (None, "")
        ]
        formula_cells = [
            worksheet_formula.cell(row=row, column=column).coordinate
            for row in range(min_row + 1, max_row + 1)
            if isinstance(worksheet_formula.cell(row=row, column=column).value, str)
            and str(worksheet_formula.cell(row=row, column=column).value).startswith("=")
        ]
        numeric_values = [
            value for value in (_safe_float(item) for item in samples) if value is not None
        ]
        lowered_header = header.lower()
        if numeric_values:
            inferred_type = "numeric"
            metric_candidates.append(header)
            if len(numeric_values) >= 2:
                avg = mean(numeric_values)
                max_value = max(numeric_values)
                min_value = min(numeric_values)
                if avg and (max_value > (avg * 1.8) or min_value < (avg * 0.2)):
                    anomalies.append(
                        {
                            "column": header,
                            "mean": round(avg, 3),
                            "min": round(min_value, 3),
                            "max": round(max_value, 3),
                        }
                    )
        elif any(token in lowered_header for token in TIME_AXIS_TOKENS):
            inferred_type = "time-axis"
            time_axis_candidates.append(header)
        else:
            inferred_type = "text"
            dimension_candidates.append(header)
        if (
            any(token in lowered_header for token in METRIC_TOKENS)
            and header not in metric_candidates
        ):
            metric_candidates.append(header)
        if (
            any(token in lowered_header for token in DIMENSION_TOKENS)
            and header not in dimension_candidates
        ):
            dimension_candidates.append(header)
        if (
            any(token in lowered_header for token in TIME_AXIS_TOKENS)
            and header not in time_axis_candidates
        ):
            time_axis_candidates.append(header)
        profile: dict[str, Any] = {
            "header": header,
            "column": get_column_letter(column),
            "inferred_type": inferred_type,
            "sample_values": [_compact_text(item, limit=48) for item in samples[:6]],
            "formula_cells": formula_cells[:12],
        }
        if numeric_values:
            profile["numeric_summary"] = {
                "count": len(numeric_values),
                "min": round(min(numeric_values), 3),
                "max": round(max(numeric_values), 3),
                "mean": round(mean(numeric_values), 3),
            }
        profiles.append(profile)
    return (
        profiles,
        deduplicate_strings(metric_candidates),
        deduplicate_strings(dimension_candidates),
        deduplicate_strings(time_axis_candidates),
        anomalies,
    )


def _openpyxl_rich_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return _sanitize_text(value)
    text = getattr(value, "tx", None)
    if text is not None:
        return _openpyxl_rich_text(text)
    rich = getattr(value, "rich", None)
    if rich is not None and hasattr(rich, "p"):
        parts: list[str] = []
        for paragraph in getattr(rich, "p", []):
            runs = getattr(paragraph, "r", []) or []
            for run in runs:
                content = _sanitize_text(getattr(run, "t", ""))
                if content:
                    parts.append(content)
        return _sanitize_text(" ".join(parts))
    str_ref = getattr(value, "strRef", None)
    if str_ref is not None:
        formula = _sanitize_text(getattr(str_ref, "f", ""))
        if formula:
            return formula
    return _sanitize_text(getattr(value, "t", "") or getattr(value, "v", ""))


def _openpyxl_formula_ref(value: Any) -> str:
    if value is None:
        return ""
    for attribute in ("f", "v"):
        content = _sanitize_text(getattr(value, attribute, ""))
        if content:
            return content
    num_ref = getattr(value, "numRef", None)
    if num_ref is not None:
        formula = _sanitize_text(getattr(num_ref, "f", ""))
        if formula:
            return formula
    str_ref = getattr(value, "strRef", None)
    if str_ref is not None:
        formula = _sanitize_text(getattr(str_ref, "f", ""))
        if formula:
            return formula
    return ""


def _range_token(value: str) -> str:
    return _sanitize_text(value).replace("$", "").replace("'", "").lower()


def _split_sheet_and_ref(value: str) -> tuple[str | None, str]:
    cleaned = _sanitize_text(value).replace("$", "")
    if "!" not in cleaned:
        return None, cleaned
    sheet_name, reference = cleaned.split("!", 1)
    return sheet_name.strip("'").lower(), reference


def _range_overlaps(reference: str, candidate: str) -> bool:
    if not reference or not candidate:
        return False
    reference_sheet, reference_range = _split_sheet_and_ref(reference)
    candidate_sheet, candidate_range = _split_sheet_and_ref(candidate)
    if reference_sheet and candidate_sheet and reference_sheet != candidate_sheet:
        return False
    try:
        ref_bounds = range_boundaries(reference_range)
        candidate_bounds = range_boundaries(candidate_range)
    except Exception:
        return _range_token(reference) in _range_token(candidate) or _range_token(
            candidate
        ) in _range_token(reference)
    return not (
        ref_bounds[2] < candidate_bounds[0]
        or candidate_bounds[2] < ref_bounds[0]
        or ref_bounds[3] < candidate_bounds[1]
        or candidate_bounds[3] < ref_bounds[1]
    )


def _chart_anchor_bbox(chart: Any, *, max_row: int, max_col: int) -> dict[str, float] | None:
    anchor = getattr(chart, "anchor", None)
    if anchor is None:
        return None
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    if start is None:
        return None
    min_col = int(getattr(start, "col", 0)) + 1
    min_row = int(getattr(start, "row", 0)) + 1
    max_col_value = int(getattr(end, "col", min_col - 1)) + 1 if end is not None else min_col + 2
    max_row_value = int(getattr(end, "row", min_row - 1)) + 1 if end is not None else min_row + 8
    return _bbox_from_grid(
        min_col=max(min_col, 1),
        min_row=max(min_row, 1),
        max_col=max(max_col_value, min_col),
        max_row=max(max_row_value, min_row),
    )


def _chart_registry(
    worksheet_value: Any,
    *,
    tabular_regions: list[dict[str, Any]],
    metric_candidates: list[str],
) -> list[dict[str, Any]]:
    charts: list[dict[str, Any]] = []
    max_row = max(int(getattr(worksheet_value, "max_row", 1) or 1), 1)
    max_col = max(int(getattr(worksheet_value, "max_column", 1) or 1), 1)
    for index, chart in enumerate(getattr(worksheet_value, "_charts", []), start=1):
        title = _compact_text(_openpyxl_rich_text(getattr(chart, "title", "")), limit=96)
        chart_type = type(chart).__name__
        bbox = _chart_anchor_bbox(chart, max_row=max_row, max_col=max_col)
        x_axis_title = _compact_text(
            _openpyxl_rich_text(getattr(getattr(chart, "x_axis", None), "title", None)),
            limit=72,
        )
        y_axis_title = _compact_text(
            _openpyxl_rich_text(getattr(getattr(chart, "y_axis", None), "title", None)),
            limit=72,
        )
        series_payload: list[dict[str, Any]] = []
        linked_refs: list[str] = []
        for series in getattr(chart, "ser", []):
            name = _compact_text(_openpyxl_rich_text(getattr(series, "tx", None)), limit=72)
            category_ref = _openpyxl_formula_ref(getattr(series, "cat", None))
            value_ref = _openpyxl_formula_ref(getattr(series, "val", None))
            if category_ref:
                linked_refs.append(category_ref)
            if value_ref:
                linked_refs.append(value_ref)
            series_payload.append(
                {
                    "name": name or f"Series {len(series_payload) + 1}",
                    "category_ref": category_ref or None,
                    "value_ref": value_ref or None,
                }
            )
        linked_table_names = [
            str(region.get("name"))
            for region in tabular_regions
            if isinstance(region.get("name"), str)
            and any(_range_overlaps(str(region.get("ref") or ""), ref) for ref in linked_refs)
        ]
        charts.append(
            {
                "ordinal": index,
                "title": title or f"Chart {index}",
                "chart_type": chart_type,
                "series_count": len(getattr(chart, "ser", [])),
                "bbox": bbox,
                "legend_present": bool(getattr(chart, "legend", None)),
                "x_axis_title": x_axis_title or None,
                "y_axis_title": y_axis_title or None,
                "series": series_payload,
                "linked_table_names": deduplicate_strings(linked_table_names),
                "metric_columns": deduplicate_strings(metric_candidates[:6]),
            }
        )
    return charts


def _pivot_registry(worksheet_formula: Any) -> list[dict[str, Any]]:
    registry: list[dict[str, Any]] = []
    for index, pivot in enumerate(getattr(worksheet_formula, "_pivots", []), start=1):
        registry.append(
            {
                "ordinal": index,
                "name": _sanitize_text(getattr(pivot, "name", "")) or f"Pivot {index}",
                "cache_id": getattr(getattr(pivot, "cache", None), "id", None),
                "data_caption": _sanitize_text(getattr(pivot, "dataCaption", "")) or None,
            }
        )
    return registry


def _drawing_anchor_bbox(
    anchor: Any,
    *,
    max_row: int,
    max_col: int,
) -> dict[str, float] | None:
    if isinstance(anchor, str) and anchor:
        try:
            min_col, min_row, max_col_ref, max_row_ref = range_boundaries(f"{anchor}:{anchor}")
        except Exception:
            return None
        return _bbox_from_grid(
            min_col=min_col,
            min_row=min_row,
            max_col=min(max_col, max_col_ref + 3),
            max_row=min(max_row, max_row_ref + 8),
        )
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    if start is None:
        return None
    min_col = int(getattr(start, "col", 0)) + 1
    min_row = int(getattr(start, "row", 0)) + 1
    max_col_value = int(getattr(end, "col", min_col - 1)) + 1 if end is not None else min_col + 3
    max_row_value = int(getattr(end, "row", min_row - 1)) + 1 if end is not None else min_row + 8
    return _bbox_from_grid(
        min_col=max(min_col, 1),
        min_row=max(min_row, 1),
        max_col=max(max_col_value, min_col),
        max_row=max(max_row_value, min_row),
    )


def _sheet_image_registry(worksheet_value: Any) -> list[dict[str, Any]]:
    images: list[dict[str, Any]] = []
    max_row = max(int(getattr(worksheet_value, "max_row", 1) or 1), 1)
    max_col = max(int(getattr(worksheet_value, "max_column", 1) or 1), 1)
    for index, image in enumerate(getattr(worksheet_value, "_images", []), start=1):
        anchor = getattr(image, "anchor", None)
        bbox = _drawing_anchor_bbox(anchor, max_row=max_row, max_col=max_col)
        name = _sanitize_text(getattr(image, "path", "")) or f"Sheet image {index}"
        images.append(
            {
                "image_ref": f"image-{index:03d}",
                "ordinal": index,
                "title": name,
                "bbox": bbox,
                "format": _sanitize_text(getattr(image, "format", "")) or None,
            }
        )
    return images


def _sheet_visual_layout(
    *,
    source_id: str,
    unit_id: str,
    render_assets: list[str],
    page_span: dict[str, int] | None,
    max_row: int,
    max_col: int,
    tabular_regions: list[dict[str, Any]],
    charts: list[dict[str, Any]],
    images: list[dict[str, Any]],
    embedded_media_assets: dict[str, str] | None,
    semantic_gap_hints: list[str],
    role_texts: list[str],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    regions: list[dict[str, Any]] = []
    artifacts: list[dict[str, Any]] = []
    counter: Counter[str] = Counter()
    for region in tabular_regions:
        ref = str(region["ref"])
        min_col, min_row, max_col_ref, max_row_ref = range_boundaries(ref)
        bbox = _bbox_from_grid(
            min_col=min_col,
            min_row=min_row,
            max_col=max_col_ref,
            max_row=max_row_ref,
        )
        counter["table"] += 1
        artifact_id = stable_artifact_id(unit_id, "table", counter["table"])
        title = f"{region['name']} table"
        payload = {
            "artifact_id": artifact_id,
            "artifact_type": "table",
            "title": title,
            "bbox": bbox,
            "normalized_bbox": normalize_bbox(bbox, width=max_col, height=max_row),
            "linked_text": ref,
            "available_channels": ["text", "render", "structure"],
            "render_assets": render_assets,
            "table_ref": ref,
        }
        regions.append(payload)
        artifacts.append(
            {
                "artifact_id": artifact_id,
                "artifact_type": "table",
                "unit_id": unit_id,
                "title": title,
                "locator_aliases": artifact_locator_aliases(
                    artifact_type="table",
                    title=title,
                    unit_title=unit_id,
                    extra_aliases=[region["name"], ref],
                ),
                "available_channels": ["text", "render", "structure"],
                "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                "graph_promoted": artifact_graph_promoted(
                    artifact_type="table", high_confidence=True
                ),
                "bbox": bbox,
                "normalized_bbox": normalize_bbox(bbox, width=max_col, height=max_row),
                "render_assets": render_assets,
                "render_page_span": page_span,
                "linked_text": ref,
                "searchable_text": "\n".join([title, ref, region["name"]]),
            }
        )
    for chart in charts:
        chart_bbox = chart.get("bbox")
        if not isinstance(chart_bbox, dict):
            continue
        counter["chart"] += 1
        artifact_id = stable_artifact_id(unit_id, "chart", counter["chart"])
        title = str(chart.get("title") or f"Chart {counter['chart']}")
        payload = {
            "artifact_id": artifact_id,
            "artifact_type": "chart",
            "title": title,
            "bbox": chart_bbox,
            "normalized_bbox": normalize_bbox(chart_bbox, width=max_col, height=max_row),
            "linked_text": title,
            "available_channels": ["text", "render", "structure"],
            "render_assets": render_assets,
            "chart_type": chart.get("chart_type"),
            "series_count": chart.get("series_count"),
            "caption_text": chart.get("title"),
            "semantic_labels": [
                value
                for value in (
                    chart.get("x_axis_title"),
                    chart.get("y_axis_title"),
                    ", ".join(
                        str(series.get("name"))
                        for series in chart.get("series", [])
                        if isinstance(series, dict) and isinstance(series.get("name"), str)
                    ),
                )
                if isinstance(value, str) and value
            ],
        }
        regions.append(payload)
        artifacts.append(
            {
                "artifact_id": artifact_id,
                "artifact_type": "chart",
                "unit_id": unit_id,
                "title": title,
                "locator_aliases": artifact_locator_aliases(
                    artifact_type="chart",
                    title=title,
                    unit_title=unit_id,
                ),
                "available_channels": ["text", "render", "structure"],
                "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                "graph_promoted": artifact_graph_promoted(
                    artifact_type="chart", high_confidence=True
                ),
                "bbox": chart_bbox,
                "normalized_bbox": normalize_bbox(chart_bbox, width=max_col, height=max_row),
                "render_assets": render_assets,
                "render_page_span": page_span,
                "linked_text": title,
                "visual_hints": [str(chart.get("chart_type") or "chart")],
                "caption_text": chart.get("title"),
                "searchable_text": "\n".join(
                    [
                        title,
                        str(chart.get("chart_type") or ""),
                        str(chart.get("x_axis_title") or ""),
                        str(chart.get("y_axis_title") or ""),
                        ", ".join(
                            str(series.get("name"))
                            for series in chart.get("series", [])
                            if isinstance(series, dict) and isinstance(series.get("name"), str)
                        ),
                        ", ".join(
                            str(value)
                            for value in chart.get("linked_table_names", [])
                            if isinstance(value, str)
                        ),
                        f"series={chart.get('series_count', 0)}",
                    ]
                ),
                "derivation_mode": "deterministic",
            }
        )
    for image in images:
        image_bbox = image.get("bbox")
        if not isinstance(image_bbox, dict):
            continue
        counter["picture"] += 1
        artifact_id = stable_artifact_id(unit_id, "picture", counter["picture"])
        title = str(image.get("title") or f"Sheet image {counter['picture']}")
        image_ref = str(image.get("image_ref") or f"image-{counter['picture']:03d}")
        media_asset = None
        if isinstance(embedded_media_assets, dict):
            media_candidate = embedded_media_assets.get(image_ref)
            if isinstance(media_candidate, str) and media_candidate:
                media_asset = media_candidate
        focus_render_assets = deduplicate_strings(
            [*([media_asset] if media_asset else []), *render_assets]
        )
        payload = {
            "artifact_id": artifact_id,
            "artifact_type": "picture",
            "title": title,
            "bbox": image_bbox,
            "normalized_bbox": normalize_bbox(image_bbox, width=max_col, height=max_row),
            "linked_text": "",
            "available_channels": ["render", "media", "structure"],
            "focus_render_assets": focus_render_assets,
            "render_assets": render_assets,
            "caption_text": None,
            "image_ref": image_ref,
            "semantic_gap_hints": ["image-heavy-sheet", "rendered-only-picture"],
        }
        regions.append(payload)
        artifacts.append(
            {
                "artifact_id": artifact_id,
                "artifact_type": "picture",
                "unit_id": unit_id,
                "title": title,
                "locator_aliases": artifact_locator_aliases(
                    artifact_type="picture",
                    title=title,
                    unit_title=unit_id,
                ),
                "available_channels": ["render", "media", "structure"],
                "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                "graph_promoted": False,
                "bbox": image_bbox,
                "normalized_bbox": normalize_bbox(image_bbox, width=max_col, height=max_row),
                "focus_render_assets": focus_render_assets,
                "render_assets": render_assets,
                "render_page_span": page_span,
                "image_ref": image_ref,
                "linked_text": "",
                "visual_hints": ["sheet-picture"],
                "searchable_text": "\n".join(
                    part
                    for part in [
                        title,
                        str(image.get("format") or ""),
                        image_ref,
                        "image-heavy-sheet",
                    ]
                    if part
                ),
                "derivation_mode": "deterministic",
                "semantic_gap_hints": ["image-heavy-sheet", "rendered-only-picture"],
            }
        )
    role_hints = _role_hints_from_text(
        role_texts,
        artifact_types=[artifact["artifact_type"] for artifact in artifacts],
    )
    return (
        {
            "artifact_type": "visual-layout",
            "source_id": source_id,
            "unit_id": unit_id,
            "render_page_span": page_span,
            "render_assets": render_assets,
            "page_size": {"width": max_col, "height": max_row},
            "role_hints": role_hints,
            "semantic_gap_hints": deduplicate_strings(semantic_gap_hints),
            "regions": regions,
        },
        artifacts,
    )


def compile_xlsx_artifacts(
    source_dir: Path,
    *,
    source_id: str,
    workbook_formula: Any,
    workbook_value: Any,
    units: list[dict[str, Any]],
    sheet_render_assets: dict[str, list[str]],
    embedded_media_assets: dict[str, dict[str, str]] | None = None,
) -> dict[str, Any]:
    spreadsheet_dir = source_dir / "spreadsheet_sheet"
    spreadsheet_dir.mkdir(parents=True, exist_ok=True)
    visual_dir = source_dir / "visual_layout"
    visual_dir.mkdir(parents=True, exist_ok=True)
    artifact_entries: list[dict[str, Any]] = []
    unit_updates: dict[str, dict[str, Any]] = {}
    visual_layout_assets: list[str] = []
    sheet_sidecars: list[str] = []
    workbook_registry = _defined_name_registry(workbook_formula)
    workbook_chart_registry: list[dict[str, Any]] = []
    workbook_formula_summary: list[dict[str, Any]] = []
    sheet_inventory: list[dict[str, Any]] = []
    formula_cross_sheet_counts: defaultdict[str, int] = defaultdict(int)
    value_lookup = {worksheet.title: worksheet for worksheet in workbook_value.worksheets}
    unit_lookup = {
        str(unit.get("title") or ""): unit for unit in units if isinstance(unit.get("title"), str)
    }
    document_render_index = 1
    for worksheet_formula in workbook_formula.worksheets:
        worksheet_value = value_lookup.get(worksheet_formula.title)
        if worksheet_value is None:
            continue
        unit = unit_lookup.get(worksheet_formula.title)
        if not isinstance(unit, dict):
            continue
        unit_id = str(unit.get("unit_id") or "")
        if not unit_id:
            continue
        render_assets = list(sheet_render_assets.get(worksheet_formula.title, []))
        page_span = (
            {"start": document_render_index, "end": document_render_index + len(render_assets) - 1}
            if render_assets
            else None
        )
        document_render_index += len(render_assets)
        formulas = _iter_formula_cells(worksheet_formula)
        for formula in formulas:
            referenced_sheets = {
                ref["sheet_name"]
                for ref in formula.get("references", [])
                if isinstance(ref, dict) and isinstance(ref.get("sheet_name"), str)
            }
            for referenced_sheet in referenced_sheets:
                if referenced_sheet != worksheet_formula.title:
                    formula_cross_sheet_counts[
                        f"{worksheet_formula.title}->{referenced_sheet}"
                    ] += 1
        regions = _tabular_regions(worksheet_value)
        primary_region_ref = str(regions[0]["ref"])
        profiles, metric_candidates, dimension_candidates, time_axis_candidates, anomalies = (
            _column_profiles(
                worksheet_value,
                worksheet_formula,
                primary_region_ref,
            )
        )
        charts = _chart_registry(
            worksheet_value,
            tabular_regions=regions,
            metric_candidates=metric_candidates,
        )
        images = _sheet_image_registry(worksheet_value)
        sheet_embedded_media = (
            embedded_media_assets.get(unit_id, {})
            if isinstance(embedded_media_assets, dict)
            else {}
        )
        workbook_chart_registry.extend(
            {
                "sheet_name": worksheet_formula.title,
                **chart,
            }
            for chart in charts
        )
        pivot_registry = _pivot_registry(worksheet_formula)
        sheet_role_hints = deduplicate_strings(
            [
                "dashboard-like"
                for _item in [1]
                if charts and metric_candidates and len(regions) <= 4
            ]
            + (["kpi-like"] if charts and metric_candidates else [])
        )
        if images:
            sheet_role_hints = deduplicate_strings(sheet_role_hints + ["image-heavy"])
        unit_gap_hints = deduplicate_strings(
            (["picture-heavy-sheet"] if images else [])
            + (["chart-table-semantic-gap"] if charts and regions else [])
        )
        sheet_payload = {
            "artifact_type": "spreadsheet-sheet",
            "source_id": source_id,
            "unit_id": unit_id,
            "sheet_name": worksheet_formula.title,
            "visibility": worksheet_formula.sheet_state,
            "used_range": _infer_primary_region(worksheet_value),
            "max_row": worksheet_value.max_row,
            "max_column": worksheet_value.max_column,
            "merged_cells": [str(item) for item in list(worksheet_value.merged_cells.ranges)[:200]],
            "freeze_panes": str(worksheet_value.freeze_panes)
            if worksheet_value.freeze_panes
            else None,
            "auto_filter": worksheet_value.auto_filter.ref if worksheet_value.auto_filter else None,
            "named_tables": [
                {"name": table["name"], "ref": table["ref"]}
                for table in regions
                if table["kind"] == "named-table"
            ],
            "tabular_regions": regions,
            "header_band": {
                "row_start": range_boundaries(primary_region_ref)[1],
                "row_end": range_boundaries(primary_region_ref)[1],
            },
            "typed_column_profiles": profiles,
            "metric_candidates": metric_candidates,
            "dimension_candidates": dimension_candidates,
            "time_axis_candidates": time_axis_candidates,
            "anomaly_scaffolds": anomalies,
            "formula_cells": formulas[:200],
            "chart_registry": charts,
            "pivot_registry": pivot_registry,
            "sheet_role_hints": sheet_role_hints,
            "semantic_gap_hints": unit_gap_hints,
            "render_assets": render_assets,
            "render_page_span": page_span,
        }
        sheet_path = spreadsheet_dir / f"{unit_id}.json"
        write_json(sheet_path, sheet_payload)
        sheet_sidecars.append(str(sheet_path.relative_to(source_dir)))
        visual_payload, visual_artifacts = _sheet_visual_layout(
            source_id=source_id,
            unit_id=unit_id,
            render_assets=render_assets,
            page_span=page_span,
            max_row=max(int(worksheet_value.max_row or 1), 1),
            max_col=max(int(worksheet_value.max_column or 1), 1),
            tabular_regions=regions,
            charts=charts,
            images=images,
            embedded_media_assets=sheet_embedded_media,
            semantic_gap_hints=unit_gap_hints,
            role_texts=[
                worksheet_formula.title,
                *metric_candidates,
                *dimension_candidates,
                *time_axis_candidates,
            ],
        )
        visual_path = visual_dir / f"{unit_id}.json"
        write_json(visual_path, visual_payload)
        visual_layout_assets.append(str(visual_path.relative_to(source_dir)))
        artifact_entries.extend(visual_artifacts)
        unit_updates[unit_id] = {
            "render_assets": render_assets,
            "render_page_span": page_span,
            "rendered_asset": render_assets[0] if render_assets else None,
            "hidden": worksheet_formula.sheet_state != "visible",
            "header_names": [profile["header"] for profile in profiles],
            "row_count": max(int(worksheet_value.max_row or 1) - 1, 0),
            "heading_aliases": [worksheet_formula.title],
            "locator_aliases": deduplicate_strings(
                [worksheet_formula.title] + metric_candidates[:3]
            ),
            "semantic_gap_hints": unit_gap_hints,
        }
        sheet_inventory.append(
            {
                "unit_id": unit_id,
                "sheet_name": worksheet_formula.title,
                "visibility": worksheet_formula.sheet_state,
                "render_page_span": page_span,
                "chart_count": len(charts),
                "table_region_count": len(regions),
                "picture_count": len(images),
                "metric_candidates": metric_candidates,
                "time_axis_candidates": time_axis_candidates,
                "sheet_role_hints": sheet_role_hints,
            }
        )
        workbook_formula_summary.append(
            {
                "sheet_name": worksheet_formula.title,
                "formula_count": len(formulas),
                "cross_sheet_reference_count": sum(
                    1
                    for formula in formulas
                    for ref in formula.get("references", [])
                    if isinstance(ref, dict)
                    and isinstance(ref.get("sheet_name"), str)
                    and ref["sheet_name"] != worksheet_formula.title
                ),
            }
        )
    workbook_payload = {
        "artifact_type": "spreadsheet-workbook",
        "schema_version": 1,
        "generated_at": _utc_now(),
        "source_id": source_id,
        "sheet_inventory": sheet_inventory,
        "named_range_registry": workbook_registry,
        "cross_sheet_reference_summary": [
            {"link": link, "count": count}
            for link, count in sorted(formula_cross_sheet_counts.items())
        ],
        "formula_dependency_summary": workbook_formula_summary,
        "chart_registry": workbook_chart_registry,
        "pivot_registry": [
            {
                "sheet_name": worksheet_formula.title,
                **pivot,
            }
            for worksheet_formula in workbook_formula.worksheets
            for pivot in _pivot_registry(worksheet_formula)
        ],
    }
    write_json(source_dir / "spreadsheet_workbook.json", workbook_payload)
    artifact_index = _write_artifact_index(
        source_dir, source_id=source_id, artifacts=artifact_entries
    )
    return {
        "artifact_index": artifact_index,
        "spreadsheet_workbook_asset": "spreadsheet_workbook.json",
        "spreadsheet_sheet_assets": sheet_sidecars,
        "visual_layout_assets": visual_layout_assets,
        "unit_updates": unit_updates,
    }


def compile_docx_visual_compatibility(
    source_dir: Path,
    *,
    source_id: str,
    units: list[dict[str, Any]],
    document_renders: list[str],
    embedded_media_assets: dict[str, str] | None = None,
) -> dict[str, Any]:
    visual_dir = source_dir / "visual_layout"
    visual_dir.mkdir(parents=True, exist_ok=True)
    artifact_entries: list[dict[str, Any]] = []
    layout_assets: list[str] = []
    unit_updates: dict[str, dict[str, Any]] = {}
    media_assets_by_ref = (
        {
            str(key): str(value)
            for key, value in embedded_media_assets.items()
            if isinstance(key, str) and key and isinstance(value, str) and value
        }
        if isinstance(embedded_media_assets, dict)
        else {}
    )
    for unit in units:
        unit_id = str(unit.get("unit_id") or "")
        if not unit_id:
            continue
        structure_asset = unit.get("structure_asset")
        structure_payload = (
            read_json(source_dir / structure_asset)
            if isinstance(structure_asset, str) and structure_asset
            else {}
        )
        blocks = [block for block in structure_payload.get("blocks", []) if isinstance(block, dict)]
        regions: list[dict[str, Any]] = []
        block_counter: Counter[str] = Counter()
        pending_caption: str | None = None
        picture_count = 0
        table_count = 0
        for block in blocks:
            text = _sanitize_text(block.get("text"))
            image_refs = [
                value for value in block.get("image_refs", []) if isinstance(value, str) and value
            ]
            if block.get("caption_kind") and text:
                pending_caption = text
                continue
            if block.get("kind") == "table":
                block_counter["table"] += 1
                table_count += 1
                artifact_id = stable_artifact_id(unit_id, "table", block_counter["table"])
                title = artifact_title_from_text(
                    text,
                    artifact_type="table",
                    fallback=f"Table {block_counter['table']}",
                )
                semantic_gap_hints = ["rendered-only-table"] if not text else []
                regions.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": "table",
                        "title": title,
                        "bbox": None,
                        "normalized_bbox": None,
                        "linked_text": text,
                        "caption_text": pending_caption,
                        "available_channels": ["text", "render", "structure"],
                        "render_assets": document_renders,
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": semantic_gap_hints,
                    }
                )
                artifact_entries.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": "table",
                        "unit_id": unit_id,
                        "title": title,
                        "locator_aliases": artifact_locator_aliases(
                            artifact_type="table",
                            title=title,
                            unit_title=str(unit.get("title") or ""),
                            extra_aliases=[pending_caption] if pending_caption else None,
                        ),
                        "available_channels": ["text", "render", "structure"],
                        "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                        "graph_promoted": artifact_graph_promoted(
                            artifact_type="table",
                            high_confidence=True,
                        ),
                        "bbox": None,
                        "normalized_bbox": None,
                        "render_assets": document_renders,
                        "render_page_span": unit.get("render_page_span"),
                        "linked_text": text,
                        "caption_text": pending_caption,
                        "searchable_text": "\n".join(
                            part for part in [title, pending_caption or "", text] if part
                        ),
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": semantic_gap_hints,
                    }
                )
                pending_caption = None
            for image_ref in image_refs:
                block_counter["picture"] += 1
                picture_count += 1
                artifact_id = stable_artifact_id(unit_id, "picture", block_counter["picture"])
                title = artifact_title_from_text(
                    pending_caption or text or image_ref,
                    artifact_type="picture",
                    fallback=f"Picture {block_counter['picture']}",
                )
                media_asset = media_assets_by_ref.get(image_ref)
                focus_render_assets = deduplicate_strings(
                    [
                        *( [media_asset] if media_asset else [] ),
                        *document_renders,
                    ]
                )
                semantic_gap_hints = (
                    [] if pending_caption or text else ["rendered-only-diagram-section"]
                )
                regions.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": "picture",
                        "title": title,
                        "bbox": None,
                        "normalized_bbox": None,
                        "linked_text": text,
                        "caption_text": pending_caption,
                        "available_channels": (
                            ["text", "render", "media", "structure"]
                            if pending_caption or text
                            else ["render", "media", "structure"]
                        ),
                        "focus_render_assets": focus_render_assets,
                        "render_assets": document_renders,
                        "image_ref": image_ref,
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": semantic_gap_hints,
                    }
                )
                artifact_entries.append(
                    {
                        "artifact_id": artifact_id,
                        "artifact_type": "picture",
                        "unit_id": unit_id,
                        "title": title,
                        "locator_aliases": artifact_locator_aliases(
                            artifact_type="picture",
                            title=title,
                            unit_title=str(unit.get("title") or ""),
                            extra_aliases=[pending_caption] if pending_caption else None,
                        ),
                        "available_channels": (
                            ["text", "render", "media", "structure"]
                            if pending_caption or text
                            else ["render", "media", "structure"]
                        ),
                        "artifact_path": str(Path("visual_layout") / f"{unit_id}.json"),
                        "graph_promoted": False,
                        "bbox": None,
                        "normalized_bbox": None,
                        "focus_render_assets": focus_render_assets,
                        "render_assets": document_renders,
                        "render_page_span": unit.get("render_page_span"),
                        "image_ref": image_ref,
                        "linked_text": text,
                        "caption_text": pending_caption,
                        "searchable_text": "\n".join(
                            part for part in [title, pending_caption or "", text, image_ref] if part
                        ),
                        "derivation_mode": "deterministic",
                        "semantic_gap_hints": semantic_gap_hints,
                    }
                )
                pending_caption = None
        unit_gap_hints = [
            value
            for value in structure_payload.get("role_hints", [])
            if isinstance(value, str) and value
        ]
        if picture_count:
            unit_gap_hints.append("image-heavy-section")
        if picture_count and not any(
            isinstance(region.get("caption_text"), str) and region.get("caption_text")
            for region in regions
            if region.get("artifact_type") == "picture"
        ):
            unit_gap_hints.append("rendered-only-diagram-section")
        if table_count and picture_count:
            unit_gap_hints.append("mixed-table-figure-section")
        layout_path = visual_dir / f"{unit_id}.json"
        write_json(
            layout_path,
            {
                "artifact_type": "visual-layout",
                "source_id": source_id,
                "unit_id": unit_id,
                "render_alignment": "document-only",
                "render_assets": document_renders,
                "role_hints": deduplicate_strings(
                    [
                        value
                        for value in structure_payload.get("role_hints", [])
                        if isinstance(value, str) and value
                    ]
                ),
                "semantic_gap_hints": deduplicate_strings(unit_gap_hints),
                "regions": regions,
            },
        )
        layout_assets.append(str(layout_path.relative_to(source_dir)))
        unit_updates[unit_id] = {
            "render_assets": document_renders,
            "semantic_gap_hints": deduplicate_strings(unit_gap_hints),
        }
    artifact_index = _write_artifact_index(
        source_dir, source_id=source_id, artifacts=artifact_entries
    )
    return {
        "artifact_index": artifact_index,
        "visual_layout_assets": layout_assets,
        "unit_updates": unit_updates,
    }
